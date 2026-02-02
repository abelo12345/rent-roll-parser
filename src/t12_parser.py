"""T12 Parser — AI-powered structure detection and GL line extraction from T12 Excel files."""

import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO

import openpyxl
import pandas as pd


@dataclass
class T12LineItem:
    source_row: int
    gl_code: str | None
    gl_description: str
    raw_label: str
    indent_level: int
    row_type: str  # leaf, section_header, subtotal, grand_total, calculated, percentage, blank
    monthly_values: list[float | None]
    total_value: float | None
    section_path: list[str] = field(default_factory=list)


@dataclass
class T12ParseResult:
    property_name: str | None
    as_of_date: datetime | None
    unit_count: int | None
    total_sf: float | None
    month_headers: list[str]
    line_items: list[T12LineItem]
    leaf_items: list[T12LineItem]
    sheet_name: str
    format_type: str  # single_col or dual_col
    month_col_start: int
    month_col_end: int
    total_col: int
    grand_totals: dict = field(default_factory=dict)  # label -> T12LineItem


def _pick_data_sheet(wb):
    """Pick the sheet with the most data rows."""
    best, best_count = wb.active, 0
    for name in wb.sheetnames:
        ws = wb[name]
        count = sum(
            1
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
            if any(c.value is not None for c in row)
        )
        if count > best_count:
            best, best_count = ws, count
    return best


def _col_letter(idx: int) -> str:
    """1-based index to Excel column letter."""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _rows_to_text(ws, max_rows: int = 40) -> str:
    """Convert first N rows of a worksheet to a text table for the LLM."""
    lines = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=False),
        start=1,
    ):
        parts = []
        for cell in row:
            col_letter = _col_letter(cell.column)
            val = cell.value
            if val is None:
                continue
            parts.append(f"{col_letter}{row_idx}={repr(val)}")
        if parts:
            lines.append(f"Row {row_idx}: " + " | ".join(parts))
    return "\n".join(lines)


def _detect_t12_format(client, ws) -> dict:
    """Use Claude to identify the T12 layout structure."""
    preview = _rows_to_text(ws, max_rows=40)

    prompt = f"""You are analyzing a Trailing 12-Month Income Statement spreadsheet. Below are the first 40 rows with cell references and values.

{preview}

Identify the structure and return ONLY valid JSON (no markdown fences, no explanation):
{{
  "format": "single_col" or "dual_col",
  "property_name": "<property name if found, else null>",
  "as_of_date": "<as-of date string e.g. '05/31/2025', else null>",
  "unit_count": <number of units if stated in headers, else null>,
  "total_sf": <total square footage if stated in headers, else null>,
  "month_header_row": <row number containing month date headers>,
  "month_col_start": <1-based column of first month>,
  "month_col_end": <1-based column of last month (12th month)>,
  "total_col": <1-based column for 'Total', else null>,
  "data_start_row": <first row with GL account data (first leaf or section header after month headers)>,
  "gl_code_col": <1-based column for GL account codes/descriptions>,
  "gl_description_col": <1-based column for separate descriptions if dual_col, else null>
}}

FORMAT DETECTION:
- "single_col": GL code and description are in the SAME column (e.g., "    40100 - Market Rent")
- "dual_col": GL code is in one column, description in a SEPARATE column (e.g., col A = "5101-0005", col B = "Gross Market Rent")

PROPERTY DETAILS:
- Look for unit count or square footage in the header area (often near the property name or as-of date)
- If not found, set to null"""

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2048,
        tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}],
        messages=[{"role": "user", "content": prompt}],
    )
    text = "".join(b.text for b in response.content if hasattr(b, "text")).strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    # Extract JSON object
    match = re.search(r"\{", text)
    if match:
        text = text[match.start():]
        depth = 0
        for i, ch in enumerate(text):
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    text = text[: i + 1]
                    break
    # Fix common AI JSON issues: trailing commas before } or ]
    text = re.sub(r",\s*([}\]])", r"\1", text)

    return json.loads(text)


def _indent_level(text: str) -> int:
    """Count leading spaces."""
    return len(text) - len(text.lstrip(" "))


# --- Row classification ---

_GRAND_TOTAL_MARKERS = [
    "total income", "total operating expenses", "net operating income",
    "total revenue", "total expenses", "net income",
]

_CALCULATED_MARKERS = [
    "gross potential", "net rental revenue", "effective gross",
    "collected (%)", "delinquent/write off",
]


def _classify_row_single_col(label: str, values: list) -> str:
    """Classify a row from single-col format T12."""
    stripped = label.strip()
    lower = stripped.lower()

    if not stripped:
        return "blank"

    if "(%)" in stripped or "(%" in stripped:
        return "percentage"

    if any(m in lower for m in _GRAND_TOTAL_MARKERS):
        return "grand_total"

    # Subtotals: starts with whitespace + "Total"
    if re.match(r"^\s+Total\s", label):
        return "subtotal"

    # Has GL account code → leaf
    if re.search(r"\d{4,5}\s*-", label):
        return "leaf"

    # No account code: section header (blank values) or calculated (has values)
    has_numeric = any(
        v is not None and str(v).strip() not in ("", " ")
        for v in values
    )
    if has_numeric:
        return "calculated"
    return "section_header"


def _classify_row_dual_col(code: str | None, desc: str | None, values: list) -> str:
    """Classify a row from dual-col format T12."""
    code_str = str(code).strip() if code else ""
    desc_str = str(desc).strip() if desc else ""
    lower_desc = desc_str.lower()

    if not code_str and not desc_str:
        return "blank"

    if "(%)" in desc_str or "(%" in desc_str:
        return "percentage"

    # Dual-col codes ending in -0000 are section headers
    if re.match(r"\d{4}-0000", code_str):
        return "section_header"

    # Codes ending in -9998 or -9999 are subtotals/grand totals
    if re.match(r"\d{4}-999[89]", code_str):
        if any(m in lower_desc for m in _GRAND_TOTAL_MARKERS):
            return "grand_total"
        return "subtotal"

    # Has a proper GL code → leaf
    if re.match(r"\d{4}-\d{4}", code_str):
        return "leaf"

    # Fallback
    has_numeric = any(
        v is not None and str(v).strip() not in ("", " ")
        for v in values
    )
    if has_numeric:
        return "calculated"
    return "section_header"


def _extract_gl_code_single(raw_label: str) -> tuple[str | None, str]:
    """Extract GL code and description from single-col label.

    Examples:
        "    40100 - Market Rent" → ("40100", "Market Rent")
        "  Market Rent"           → (None, "Market Rent")
    """
    m = re.search(r"(\d{4,5})\s*-\s*(.+)", raw_label)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, raw_label.strip()


def _extract_all_line_items(ws, fmt: dict) -> list[T12LineItem]:
    """Extract all line items from the T12 worksheet."""
    format_type = fmt["format"]
    data_start = fmt["data_start_row"]
    month_start = fmt["month_col_start"]
    month_end = fmt["month_col_end"]
    total_col = fmt.get("total_col")
    gl_code_col = fmt["gl_code_col"]
    gl_desc_col = fmt.get("gl_description_col")

    num_months = month_end - month_start + 1

    items = []
    section_stack = []  # track current section hierarchy

    for row_idx in range(data_start, ws.max_row + 1):
        # Read monthly values
        monthly = []
        for col in range(month_start, month_end + 1):
            v = ws.cell(row=row_idx, column=col).value
            if v is not None:
                try:
                    monthly.append(float(v))
                except (ValueError, TypeError):
                    monthly.append(None)
            else:
                monthly.append(None)

        # Pad to 12 if fewer months
        while len(monthly) < 12:
            monthly.append(None)

        # Read total
        total_val = None
        if total_col:
            tv = ws.cell(row=row_idx, column=total_col).value
            if tv is not None:
                try:
                    total_val = float(tv)
                except (ValueError, TypeError):
                    total_val = None

        # Build label and classify
        if format_type == "dual_col" and gl_desc_col:
            code_raw = ws.cell(row=row_idx, column=gl_code_col).value
            desc_raw = ws.cell(row=row_idx, column=gl_desc_col).value
            raw_label = f"{code_raw or ''} {desc_raw or ''}".strip()
            code_str = str(code_raw).strip() if code_raw else None
            desc_str = str(desc_raw).strip() if desc_raw else ""
            gl_code = code_str
            gl_description = desc_str
            indent = _indent_level(str(desc_raw)) if desc_raw else 0
            row_type = _classify_row_dual_col(code_raw, desc_raw, monthly)
        else:
            cell_val = ws.cell(row=row_idx, column=gl_code_col).value
            raw_label = str(cell_val) if cell_val is not None else ""
            indent = _indent_level(raw_label)
            gl_code, gl_description = _extract_gl_code_single(raw_label)
            row_type = _classify_row_single_col(raw_label, monthly)

        # Update section stack
        if row_type == "section_header":
            # Pop deeper or equal levels
            while section_stack and section_stack[-1][1] >= indent:
                section_stack.pop()
            section_stack.append((gl_description or raw_label.strip(), indent))
        elif row_type == "subtotal":
            # Pop one level
            if section_stack:
                section_stack.pop()

        item = T12LineItem(
            source_row=row_idx,
            gl_code=gl_code,
            gl_description=gl_description,
            raw_label=raw_label,
            indent_level=indent,
            row_type=row_type,
            monthly_values=monthly[:12],
            total_value=total_val,
            section_path=[s[0] for s in section_stack],
        )
        items.append(item)

    return items


def _dedup_leaf_items(items: list[T12LineItem]) -> list[T12LineItem]:
    """Remove duplicate leaf rows with the same GL code and identical totals.

    Checks both within same section AND across sections (same code + same total).
    Keeps the more deeply indented (more specific) version.
    """
    # First pass: find all leaf items by GL code
    code_items: dict[str, list[T12LineItem]] = {}
    for item in items:
        if item.row_type == "leaf" and item.gl_code:
            code_items.setdefault(item.gl_code, []).append(item)

    # Mark duplicates: same code, same total, keep the deeper-indented one
    for code, code_list in code_items.items():
        if len(code_list) < 2:
            continue
        for i in range(len(code_list)):
            for j in range(i + 1, len(code_list)):
                a, b = code_list[i], code_list[j]
                if (
                    a.row_type == "leaf"
                    and b.row_type == "leaf"
                    and a.total_value is not None
                    and b.total_value is not None
                    and abs(a.total_value - b.total_value) < 1
                ):
                    # Keep the deeper-indented one
                    if a.indent_level >= b.indent_level:
                        b.row_type = "duplicate"
                    else:
                        a.row_type = "duplicate"

    return items


def _parse_as_of_date(date_str: str | None) -> datetime | None:
    """Try to parse an as-of date string."""
    if not date_str:
        return None
    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    # Try extracting date from longer strings like "12 months ending February 28, 2025"
    m = re.search(r"(\w+ \d{1,2},?\s*\d{4})", str(date_str))
    if m:
        for fmt in ["%B %d, %Y", "%B %d %Y", "%b %d, %Y"]:
            try:
                return datetime.strptime(m.group(1).replace(",", ", ").strip(), fmt)
            except ValueError:
                continue
    return None


def _extract_grand_totals(items: list[T12LineItem]) -> dict:
    """Extract grand total rows keyed by normalized label."""
    totals = {}
    for item in items:
        if item.row_type == "grand_total":
            lower = item.gl_description.lower().strip()
            if "total income" in lower or "total revenue" in lower:
                totals["total_income"] = item
            elif "total operating expenses" in lower or "total expenses" in lower:
                totals["total_opex"] = item
            elif "net operating income" in lower or "net income" in lower:
                # Prefer NOI over net income if both exist
                key = "noi" if "operating" in lower else "net_income"
                totals[key] = item
    return totals


def parse_t12(client, file_bytes: bytes, filename: str) -> T12ParseResult:
    """Parse a T12 Excel file. Returns T12ParseResult."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = _pick_data_sheet(wb)
    sheet_name = ws.title

    # AI-detect format
    fmt = _detect_t12_format(client, ws)

    # Defaults
    fmt.setdefault("format", "single_col")
    fmt.setdefault("total_col", None)
    fmt.setdefault("gl_description_col", None)

    # Extract all line items
    all_items = _extract_all_line_items(ws, fmt)

    # Dedup
    all_items = _dedup_leaf_items(all_items)

    # Separate leaves
    leaf_items = [item for item in all_items if item.row_type == "leaf"]

    # Extract grand totals for checks
    grand_totals = _extract_grand_totals(all_items)

    # Parse as-of date
    as_of = _parse_as_of_date(fmt.get("as_of_date"))

    # Month headers
    month_headers = []
    header_row = fmt.get("month_header_row")
    if header_row:
        for col in range(fmt["month_col_start"], fmt["month_col_end"] + 1):
            v = ws.cell(row=header_row, column=col).value
            month_headers.append(str(v) if v else "")

    return T12ParseResult(
        property_name=fmt.get("property_name"),
        as_of_date=as_of,
        unit_count=fmt.get("unit_count"),
        total_sf=fmt.get("total_sf"),
        month_headers=month_headers,
        line_items=all_items,
        leaf_items=leaf_items,
        sheet_name=sheet_name,
        format_type=fmt.get("format", "single_col"),
        month_col_start=fmt["month_col_start"],
        month_col_end=fmt["month_col_end"],
        total_col=fmt.get("total_col") or (fmt["month_col_end"] + 1),
        grand_totals=grand_totals,
    )
