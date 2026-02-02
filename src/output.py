"""Step 4: Write multi-sheet Excel with formulas & formatting matching RR Output.xlsx."""

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# ---------------------------------------------------------------------------
# Style constants (matching RR Output.xlsx)
# ---------------------------------------------------------------------------
FONT_NAME = "Garamond"
FONT_SIZE = 10
TITLE_FILL = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
TITLE_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")
HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
DATA_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
BOLD_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)

THIN_SIDE = Side(style="thin")
HAIR_SIDE = Side(style="hair")
NO_SIDE = Side()

CENTER_CONT = Alignment(horizontal="centerContinuous", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

# Number formats
FMT_PCT = "0.0%"
FMT_CURRENCY = '"$"#,##0_);\\("$"#,##0\\)'
FMT_CURRENCY_DEC = '"$"#,##0.00_);\\("$"#,##0.00\\)'
FMT_NUMBER = "#,##0"

# Section divider columns (right border = hair) — F, L, R, X, AD
SECTION_DIVIDER_COLS = [6, 12, 18, 24, 30]  # 1-based column indices for F, L, R, X, AD

# Column format mapping for each section position (1-indexed within section)
# Sections: Unit Mix (B-F: 4 data cols), then 5 sections of 6 data cols each
# Unit Mix: Units, Avg SF, Total SF, % of Total
# Others: Units, Avg SF, Total SF, % of Total, $/Unit, $/SF
SECTION_FORMATS = {
    "units": None,  # General
    "avg_sf": FMT_NUMBER,
    "total_sf": FMT_NUMBER,
    "pct": FMT_PCT,
    "per_unit": FMT_CURRENCY,
    "per_sf": FMT_CURRENCY_DEC,
}

# Map column index to format (for cols C through AJ)
COL_FORMATS = {}
# Unit Mix: C=units, D=avg_sf, E=total_sf, F=pct
for i, fmt_key in enumerate(["units", "avg_sf", "total_sf", "pct"]):
    COL_FORMATS[3 + i] = SECTION_FORMATS[fmt_key]  # C=3, D=4, E=5, F=6
# Sections starting at G, M, S, Y, AE (cols 7, 13, 19, 25, 31)
for section_start in [7, 13, 19, 25, 31]:
    for i, fmt_key in enumerate(["units", "avg_sf", "total_sf", "pct", "per_unit", "per_sf"]):
        COL_FORMATS[section_start + i] = SECTION_FORMATS[fmt_key]


# Standardized sheet column layout (base columns before charge columns)
STD_COLUMNS_BASE = [
    ("A", "Unit"),
    ("B", "Floorplan"),
    ("C", "Unit Type"),
    ("D", "Display Type"),
    ("E", "Reno"),
    ("F", "SQFT"),
    ("G", "Status"),
    ("H", "Occupancy"),
    ("I", "Tenant"),
    ("J", "Move-In"),
    ("K", "Move-Out"),
    ("L", "Lease Start"),
    ("M", "Lease End"),
    ("N", "Market Rent"),
    ("O", "Lease Rent"),
    ("P", "Total Billing"),
]

# Default layout when no charge columns (backward compatible)
STD_COLUMNS = STD_COLUMNS_BASE + [("Q", "Dedup Flag"), ("R", "Source Row")]


def build_std_columns(charge_col_names: list[str]) -> tuple[list[tuple[str, str]], dict[str, str]]:
    """Build the full Standardized column layout with charge columns inserted.

    Returns (full_columns_list, charge_std_map) where charge_std_map maps
    charge name -> Standardized column letter.
    """
    cols = list(STD_COLUMNS_BASE)  # A through P
    next_col_idx = 17  # Q is 17 (1-based), first slot after P
    charge_std_map = {}

    for name in charge_col_names:
        letter = get_column_letter(next_col_idx)
        cols.append((letter, name))
        charge_std_map[name] = letter
        next_col_idx += 1

    # Dedup Flag and Source Row go after all charge columns
    dedup_letter = get_column_letter(next_col_idx)
    source_letter = get_column_letter(next_col_idx + 1)
    cols.append((dedup_letter, "Dedup Flag"))
    cols.append((source_letter, "Source Row"))

    return cols, charge_std_map

# Fields in the standardized sheet that reference raw data columns
RAW_REF_FIELDS = {
    "unit": "A",
    "floorplan": "B",
    "sqft": "F",
    "status": "G",
    "tenant_name": "I",
    "move_in": "J",
    "move_out": "K",
    "lease_start": "L",
    "lease_end": "M",
    "market_rent": "N",
    "lease_rent": "O",
    "total_billing": "P",
}


def _apply_border(ws, row, min_col, max_col, top=None, bottom=None, left_col=None, right_col=None):
    """Apply borders to a range of cells."""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        existing = cell.border
        new_left = existing.left
        new_right = existing.right
        new_top = top if top else existing.top
        new_bottom = bottom if bottom else existing.bottom

        if col == left_col:
            new_left = THIN_SIDE
        if col == right_col:
            new_right = THIN_SIDE
        if col in SECTION_DIVIDER_COLS:
            new_right = HAIR_SIDE

        cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)


def write_raw_data_sheet(wb: openpyxl.Workbook, raw_wb_bytes: bytes, sheet_name: str):
    """Write the 'Raw Data' sheet — verbatim copy of uploaded rent roll."""
    raw_wb = openpyxl.load_workbook(BytesIO(raw_wb_bytes), data_only=True)
    raw_ws = raw_wb[sheet_name]

    ws = wb.create_sheet("Raw Data")
    for row in raw_ws.iter_rows(min_row=1, max_row=raw_ws.max_row, max_col=raw_ws.max_column):
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            # Copy basic formatting
            if cell.number_format:
                new_cell.number_format = cell.number_format

    # Copy merged cells
    for merge in raw_ws.merged_cells.ranges:
        ws.merge_cells(str(merge))

    return ws


def write_standardized_sheet(wb: openpyxl.Workbook, agg_data: dict, column_map: dict):
    """Write the 'Standardized' sheet with formulas referencing Raw Data.

    Returns (ws, std_columns, charge_std_map) so downstream sheets know the layout.
    """
    ws = wb.create_sheet("Standardized")
    full_df = agg_data["df"]
    cols = column_map["columns"]
    charge_cols = column_map.get("charge_columns") or {}
    is_multi_row = column_map.get("format") == "multi_row"
    status_col_exists = column_map.get("status_column_exists", True)

    # For multi-row format, gather charge column names from the DataFrame
    # since they come from row-based charge codes, not column headers
    if is_multi_row:
        extra_charge_names = sorted(
            c.replace("charge_", "") for c in full_df.columns if c.startswith("charge_")
        )
        charge_names = extra_charge_names
    else:
        charge_names = list(charge_cols.keys())

    std_columns, charge_std_map = build_std_columns(charge_names)

    # Figure out Dedup Flag and Source Row column indices from the built layout
    dedup_col_idx = None
    source_col_idx = None
    for letter, header in std_columns:
        idx = openpyxl.utils.column_index_from_string(letter)
        if header == "Dedup Flag":
            dedup_col_idx = idx
        elif header == "Source Row":
            source_col_idx = idx

    # Write header row
    for col_letter, header in std_columns:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = BOLD_FONT

    # Write data rows
    for i, (_, row) in enumerate(full_df.iterrows()):
        excel_row = i + 2  # 1-based, skip header
        source_row = int(row["_source_row"])

        # Unit — formula to Raw Data
        unit_col = cols.get("unit")
        if unit_col:
            ws.cell(row=excel_row, column=1).value = f"='Raw Data'!{unit_col}{source_row}"

        # Floorplan — formula to Raw Data
        fp_col = cols.get("floorplan")
        if fp_col:
            ws.cell(row=excel_row, column=2).value = f"='Raw Data'!{fp_col}{source_row}"

        # Unit Type — static from mapping
        ws.cell(row=excel_row, column=3, value=row.get("unit_type", ""))
        # Display Type — static
        ws.cell(row=excel_row, column=4, value=row.get("display_type", ""))
        # Reno — static
        ws.cell(row=excel_row, column=5, value="Y" if row.get("reno") else "N")

        # SQFT — formula to Raw Data
        sqft_col = cols.get("sqft")
        if sqft_col:
            ws.cell(row=excel_row, column=6).value = f"='Raw Data'!{sqft_col}{source_row}"

        # Status — formula to Raw Data if column exists, otherwise static
        status_col = cols.get("status")
        if status_col and status_col_exists:
            ws.cell(row=excel_row, column=7).value = f"='Raw Data'!{status_col}{source_row}"
        else:
            ws.cell(row=excel_row, column=7, value=row.get("status", ""))

        # Occupancy — static classification
        ws.cell(row=excel_row, column=8, value=row.get("occupancy", ""))

        # Tenant — formula to Raw Data
        tenant_col = cols.get("tenant_name")
        if tenant_col:
            ws.cell(row=excel_row, column=9).value = f"='Raw Data'!{tenant_col}{source_row}"

        # Date fields — formulas to Raw Data, with text-to-date conversion
        for field, std_col in [("move_in", 10), ("move_out", 11), ("lease_start", 12), ("lease_end", 13)]:
            raw_col = cols.get(field)
            if raw_col:
                ref = f"'Raw Data'!{raw_col}{source_row}"
                formula = (
                    f'=IF({ref}="","",IF(ISNUMBER({ref}),{ref},DATEVALUE({ref})))'
                )
                ws.cell(row=excel_row, column=std_col).value = formula
                ws.cell(row=excel_row, column=std_col).number_format = "MM/DD/YYYY"

        # Market Rent — formula to Raw Data if column exists, static otherwise
        mr_col = cols.get("market_rent")
        if mr_col:
            ws.cell(row=excel_row, column=14).value = f"='Raw Data'!{mr_col}{source_row}"
            ws.cell(row=excel_row, column=14).number_format = FMT_CURRENCY
        else:
            val = row.get("market_rent")
            if val is not None and val == val:  # not NaN
                ws.cell(row=excel_row, column=14, value=val)
                ws.cell(row=excel_row, column=14).number_format = FMT_CURRENCY

        # Lease Rent — formula for single-row, static for multi-row (aggregated from charge rows)
        lr_col = cols.get("lease_rent")
        if lr_col and not is_multi_row:
            ws.cell(row=excel_row, column=15).value = f"='Raw Data'!{lr_col}{source_row}"
            ws.cell(row=excel_row, column=15).number_format = FMT_CURRENCY
        else:
            val = row.get("lease_rent")
            if val is not None and val == val:  # not NaN
                ws.cell(row=excel_row, column=15, value=val)
                ws.cell(row=excel_row, column=15).number_format = FMT_CURRENCY

        # Total Billing — formula for single-row, static for multi-row
        tb_col = cols.get("total_billing")
        if tb_col and not is_multi_row:
            ws.cell(row=excel_row, column=16).value = f"='Raw Data'!{tb_col}{source_row}"
            ws.cell(row=excel_row, column=16).number_format = FMT_CURRENCY
        else:
            val = row.get("total_billing")
            if val is not None and val == val:  # not NaN
                ws.cell(row=excel_row, column=16, value=val)
                ws.cell(row=excel_row, column=16).number_format = FMT_CURRENCY

        # Charge columns
        if is_multi_row:
            # Multi-row: charges are static values from parsed records
            for charge_name in charge_names:
                if charge_name in charge_std_map:
                    std_letter = charge_std_map[charge_name]
                    std_col_idx = openpyxl.utils.column_index_from_string(std_letter)
                    val = row.get(f"charge_{charge_name}")
                    if val is not None and val == val:  # not NaN
                        ws.cell(row=excel_row, column=std_col_idx, value=val)
                        ws.cell(row=excel_row, column=std_col_idx).number_format = FMT_CURRENCY
        else:
            # Single-row: formulas to Raw Data
            for charge_name, raw_letter in charge_cols.items():
                if raw_letter and charge_name in charge_std_map:
                    std_letter = charge_std_map[charge_name]
                    std_col_idx = openpyxl.utils.column_index_from_string(std_letter)
                    ws.cell(row=excel_row, column=std_col_idx).value = (
                        f"='Raw Data'!{raw_letter}{source_row}"
                    )
                    ws.cell(row=excel_row, column=std_col_idx).number_format = FMT_CURRENCY

        # Dedup Flag
        ws.cell(row=excel_row, column=dedup_col_idx,
                value="DROPPED" if row["_dedup_flag"] else "")

        # Source Row
        ws.cell(row=excel_row, column=source_col_idx, value=source_row)

    # Auto-fit column widths (approximate)
    for col_letter, _ in std_columns:
        ws.column_dimensions[col_letter].width = 14

    return ws, std_columns, charge_std_map


def write_summary_sheet(wb: openpyxl.Workbook, agg_data: dict, date_ranges: list[dict],
                        dedup_col: str = "Q"):
    """Write the 'Summary' sheet with COUNTIFS/SUMIFS/AVERAGEIFS formulas.

    date_ranges: list of {"label": str, "start": datetime, "end": datetime}
    dedup_col: column letter for the Dedup Flag in Standardized sheet
    """
    ws = wb.create_sheet("Summary")

    unit_type_order = agg_data["unit_type_order"]
    active_df = agg_data["active_df"]
    total_rows = len(agg_data["df"])  # Including deduped, for Standardized row count
    std_data_rows = total_rows  # Total rows in Standardized (including deduped)
    std_last_row = std_data_rows + 1  # +1 for header

    # Separate non-reno and reno types
    non_reno_types = [t for t in unit_type_order if not t.endswith(" Reno")]
    reno_types = [t for t in unit_type_order if t.endswith(" Reno")]

    # Build row layout: row 6 starts non-reno, blank separator, then reno, blank separator, totals
    data_rows = []
    for t in non_reno_types:
        data_rows.append(t)
    # Row 13 equivalent = blank separator
    data_rows.append(None)  # separator
    for t in reno_types:
        data_rows.append(t)
    data_rows.append(None)  # separator before totals

    first_data_row = 6
    total_row = first_data_row + len(data_rows)

    # Sections definition
    # Each section: (start_col_letter, label, has_dollar_cols, date_filter)
    sections = [
        ("C", "Unit Mix", False, None),
        ("G", "Occupied Units", True, None),
        ("M", "Market Rents", True, None),
    ]
    for dr in date_ranges:
        sections.append((None, dr["label"], True, dr))

    # Assign column letters to sections
    current_col = 3  # Start at C
    section_defs = []
    for _, label, has_dollar, date_filter in sections:
        start_col = current_col
        if has_dollar:
            ncols = 6  # Units, Avg SF, Total SF, % of Total, $/Unit, $/SF
        else:
            ncols = 4  # Units, Avg SF, Total SF, % of Total
        section_defs.append({
            "start_col": start_col,
            "ncols": ncols,
            "label": label,
            "has_dollar": has_dollar,
            "date_filter": date_filter,
        })
        current_col += ncols

    last_col = current_col - 1  # Last data column (1-based)

    # --- Row 2: Title banner ---
    ws.cell(row=2, column=2, value="Unit Mix & Pricing Summary")
    for col in range(2, last_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.alignment = CENTER_CONT

    # --- Row 3: Section headers ---
    for sdef in section_defs:
        cell = ws.cell(row=3, column=sdef["start_col"], value=sdef["label"])
        cell.font = HEADER_FONT
        cell.alignment = CENTER_CONT
        # Apply center-continuous to all cells in section
        for c in range(sdef["start_col"], sdef["start_col"] + sdef["ncols"]):
            ws.cell(row=3, column=c).alignment = CENTER_CONT
            ws.cell(row=3, column=c).font = HEADER_FONT

    # --- Row 4: Sub-headers ---
    ws.cell(row=4, column=2, value="Unit Type").font = HEADER_FONT
    ws.cell(row=4, column=2).alignment = RIGHT_ALIGN
    for sdef in section_defs:
        col = sdef["start_col"]
        headers = ["Units", "Avg. SF", "Total SF", "% of Total"]
        if sdef["has_dollar"]:
            headers += ["$ / Unit", "$ / SF"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=4, column=col + i, value=h)
            cell.font = HEADER_FONT
            cell.alignment = RIGHT_ALIGN

    # --- Row 5: Empty ---

    # --- Data rows ---
    # Standardized sheet references:
    # D = Display Type, F = SQFT, H = Occupancy, L = Lease Start, M = Lease End
    # N = Market Rent, O = Lease Rent, Q = Dedup Flag
    # We use the full column range: 2 to std_last_row

    for i, type_name in enumerate(data_rows):
        row = first_data_row + i
        if type_name is None:
            continue  # blank separator row

        ws.cell(row=row, column=2, value=type_name).font = DATA_FONT

        for sdef in section_defs:
            col = sdef["start_col"]
            _write_section_formulas(
                ws, row, col, type_name, sdef, std_last_row, total_row,
                first_data_row, data_rows, dedup_col
            )

    # --- Total / WA row ---
    ws.cell(row=total_row, column=2, value="Total / WA").font = BOLD_FONT

    # Collect all data row numbers (non-blank) for SUM formulas
    data_row_numbers = [first_data_row + i for i, t in enumerate(data_rows) if t is not None]

    for sdef in section_defs:
        col = sdef["start_col"]
        _write_total_row_formulas(ws, total_row, col, sdef, data_row_numbers, std_last_row,
                                  dedup_col)

    # --- Apply formatting ---
    _apply_all_formatting(ws, first_data_row, data_rows, total_row, last_col, section_defs)

    return ws


def _write_section_formulas(ws, row, col, type_name, sdef, std_last_row, total_row,
                            first_data_row, data_rows, dedup_col="Q"):
    """Write formulas for one section (6 or 4 columns) for a specific unit type row."""
    # Common criteria: display type match + not deduped
    type_crit = f'Standardized!$D$2:$D${std_last_row},"{type_name}"'
    dedup_crit = f'Standardized!${dedup_col}$2:${dedup_col}${std_last_row},""'

    # Additional criteria for section type
    if sdef["label"] == "Unit Mix":
        # All units of this type (not deduped)
        count_criteria = f"{type_crit},{dedup_crit}"
        sf_criteria = count_criteria
        rent_col = None
    elif sdef["label"] == "Occupied Units":
        # Occupied units: Occupancy = "Occupied"
        occ_crit = f'Standardized!$H$2:$H${std_last_row},"Occupied"'
        count_criteria = f"{type_crit},{dedup_crit},{occ_crit}"
        sf_criteria = count_criteria
        rent_col = "$O"  # Lease Rent
    elif sdef["label"] == "Market Rents":
        # All units of this type
        count_criteria = f"{type_crit},{dedup_crit}"
        sf_criteria = count_criteria
        rent_col = "$N"  # Market Rent
    else:
        # Date-filtered section: occupied units with move-in date within range
        df = sdef["date_filter"]
        # Move-In ($J) >= start AND Move-In ($J) <= end
        occ_crit = f'Standardized!$H$2:$H${std_last_row},"Occupied"'
        movein_gte = f'Standardized!$J$2:$J${std_last_row},">=" & DATE({df["start"].year},{df["start"].month},{df["start"].day})'
        movein_lte = f'Standardized!$J$2:$J${std_last_row},"<=" & DATE({df["end"].year},{df["end"].month},{df["end"].day})'
        count_criteria = f"{type_crit},{dedup_crit},{occ_crit},{movein_gte},{movein_lte}"
        sf_criteria = count_criteria
        rent_col = "$O"  # Lease Rent

    sf_range = f"Standardized!$F$2:$F${std_last_row}"
    count_range = f"Standardized!$D$2:$D${std_last_row}"

    # Units = COUNTIFS(...)
    units_formula = f"=COUNTIFS({count_criteria})"
    cell = ws.cell(row=row, column=col, value=units_formula)

    # Avg. SF = AVERAGEIFS(sqft_range, criteria...)
    avg_sf_formula = f"=IFERROR(AVERAGEIFS({sf_range},{count_criteria}),0)"
    ws.cell(row=row, column=col + 1, value=avg_sf_formula)

    # Total SF = SUMIFS(sqft_range, criteria...)
    total_sf_formula = f"=SUMIFS({sf_range},{count_criteria})"
    ws.cell(row=row, column=col + 2, value=total_sf_formula)

    # % of Total = Total SF / Total SF in total row
    total_sf_col_letter = get_column_letter(col + 2)
    total_row_ref = f"${total_sf_col_letter}${total_row}"
    pct_formula = f"=IFERROR({total_sf_col_letter}{row}/{total_row_ref},0)"
    ws.cell(row=row, column=col + 3, value=pct_formula)

    if sdef["has_dollar"] and rent_col:
        rent_range = f"Standardized!{rent_col}$2:{rent_col}${std_last_row}"
        units_col_letter = get_column_letter(col)

        # $ / Unit = SUMIFS(rent, criteria...) / COUNTIFS(criteria...)
        per_unit_formula = f"=IFERROR(SUMIFS({rent_range},{count_criteria})/{units_col_letter}{row},0)"
        ws.cell(row=row, column=col + 4, value=per_unit_formula)

        # $ / SF = SUMIFS(rent, criteria...) / SUMIFS(sqft, criteria...)
        total_sf_cell = f"{total_sf_col_letter}{row}"
        per_sf_formula = f"=IFERROR(SUMIFS({rent_range},{count_criteria})/{total_sf_cell},0)"
        ws.cell(row=row, column=col + 5, value=per_sf_formula)
    elif sdef["has_dollar"]:
        # Market rents with no specific rent column handled above
        pass


def _write_total_row_formulas(ws, total_row, col, sdef, data_row_numbers, std_last_row,
                              dedup_col="Q"):
    """Write total/weighted-average formulas for the totals row."""
    # Units = SUM of all unit type rows
    units_col = get_column_letter(col)
    units_refs = "+".join(f"{units_col}{r}" for r in data_row_numbers)
    ws.cell(row=total_row, column=col, value=f"={units_refs}").font = BOLD_FONT

    # Total SF = SUM of all unit type rows
    sf_col = get_column_letter(col + 2)
    sf_refs = "+".join(f"{sf_col}{r}" for r in data_row_numbers)
    ws.cell(row=total_row, column=col + 2, value=f"={sf_refs}").font = BOLD_FONT

    # Avg. SF = Total SF / Units (weighted average)
    avg_sf_col = get_column_letter(col + 1)
    ws.cell(
        row=total_row, column=col + 1,
        value=f"=IFERROR({sf_col}{total_row}/{units_col}{total_row},0)"
    ).font = BOLD_FONT

    # % of Total = should be 100% (or sum)
    pct_col = get_column_letter(col + 3)
    pct_refs = "+".join(f"{pct_col}{r}" for r in data_row_numbers)
    ws.cell(row=total_row, column=col + 3, value=f"={pct_refs}").font = BOLD_FONT

    if sdef["has_dollar"]:
        # $ / Unit (WA) = total rent / total units
        per_unit_col = get_column_letter(col + 4)
        # Build sum of ($/unit * units) for each row, then divide by total units
        # Simpler: reference Standardized directly
        if sdef["label"] == "Occupied Units":
            rent_col = "$O"
            occ_crit = f'Standardized!$H$2:$H${std_last_row},"Occupied"'
            dedup_crit = f'Standardized!${dedup_col}$2:${dedup_col}${std_last_row},""'
            rent_range = f"Standardized!{rent_col}$2:{rent_col}${std_last_row}"
            ws.cell(
                row=total_row, column=col + 4,
                value=f"=IFERROR(SUMIFS({rent_range},{occ_crit},{dedup_crit})/{units_col}{total_row},0)"
            ).font = BOLD_FONT
            sf_range = f"Standardized!$F$2:$F${std_last_row}"
            ws.cell(
                row=total_row, column=col + 5,
                value=f"=IFERROR(SUMIFS({rent_range},{occ_crit},{dedup_crit})/SUMIFS({sf_range},{occ_crit},{dedup_crit}),0)"
            ).font = BOLD_FONT
        elif sdef["label"] == "Market Rents":
            rent_col = "$N"
            dedup_crit = f'Standardized!${dedup_col}$2:${dedup_col}${std_last_row},""'
            rent_range = f"Standardized!{rent_col}$2:{rent_col}${std_last_row}"
            ws.cell(
                row=total_row, column=col + 4,
                value=f"=IFERROR(SUMIFS({rent_range},{dedup_crit})/{units_col}{total_row},0)"
            ).font = BOLD_FONT
            sf_range = f"Standardized!$F$2:$F${std_last_row}"
            ws.cell(
                row=total_row, column=col + 5,
                value=f"=IFERROR(SUMIFS({rent_range},{dedup_crit})/SUMIFS({sf_range},{dedup_crit}),0)"
            ).font = BOLD_FONT
        else:
            # Date-filtered: sum of rents / count for occupied in move-in date range
            df = sdef["date_filter"]
            rent_col = "$O"
            occ_crit = f'Standardized!$H$2:$H${std_last_row},"Occupied"'
            dedup_crit = f'Standardized!${dedup_col}$2:${dedup_col}${std_last_row},""'
            movein_gte = f'Standardized!$J$2:$J${std_last_row},">=" & DATE({df["start"].year},{df["start"].month},{df["start"].day})'
            movein_lte = f'Standardized!$J$2:$J${std_last_row},"<=" & DATE({df["end"].year},{df["end"].month},{df["end"].day})'
            all_crit = f"{occ_crit},{dedup_crit},{movein_gte},{movein_lte}"
            rent_range = f"Standardized!{rent_col}$2:{rent_col}${std_last_row}"
            sf_range = f"Standardized!$F$2:$F${std_last_row}"
            ws.cell(
                row=total_row, column=col + 4,
                value=f"=IFERROR(SUMIFS({rent_range},{all_crit})/{units_col}{total_row},0)"
            ).font = BOLD_FONT
            ws.cell(
                row=total_row, column=col + 5,
                value=f"=IFERROR(SUMIFS({rent_range},{all_crit})/SUMIFS({sf_range},{all_crit}),0)"
            ).font = BOLD_FONT


def _apply_all_formatting(ws, first_data_row, data_rows, total_row, last_col, section_defs):
    """Apply all formatting: fonts, number formats, borders, column widths."""
    # Recalculate section divider columns based on actual sections
    divider_cols = set()
    for sdef in section_defs:
        last_section_col = sdef["start_col"] + sdef["ncols"] - 1
        divider_cols.add(last_section_col)

    # Number formats for all data + total rows
    all_rows = list(range(first_data_row, total_row + 1))
    for sdef in section_defs:
        col = sdef["start_col"]
        formats = [None, FMT_NUMBER, FMT_NUMBER, FMT_PCT]
        if sdef["has_dollar"]:
            formats += [FMT_CURRENCY, FMT_CURRENCY_DEC]
        for i, fmt in enumerate(formats):
            if fmt:
                for r in all_rows:
                    ws.cell(row=r, column=col + i).number_format = fmt

    # Font for data rows
    for i, type_name in enumerate(data_rows):
        row = first_data_row + i
        if type_name is None:
            continue
        for c in range(2, last_col + 1):
            cell = ws.cell(row=row, column=c)
            if cell.font == BOLD_FONT:
                continue
            cell.font = DATA_FONT

    # Borders
    # Row 2: top + bottom thin, left on B, right on last_col
    for c in range(2, last_col + 1):
        cell = ws.cell(row=2, column=c)
        left = THIN_SIDE if c == 2 else NO_SIDE
        right = THIN_SIDE if c == last_col else NO_SIDE
        cell.border = Border(top=THIN_SIDE, bottom=THIN_SIDE, left=left, right=right)

    # Rows 3-4: left on B, right on last_col, section dividers
    for r in [3, 4]:
        for c in range(2, last_col + 1):
            cell = ws.cell(row=r, column=c)
            left = THIN_SIDE if c == 2 else NO_SIDE
            right = THIN_SIDE if c == last_col else (HAIR_SIDE if c in divider_cols else NO_SIDE)
            top = THIN_SIDE if (r == 3 and c in divider_cols) else NO_SIDE
            cell.border = Border(top=top, left=left, right=right)

    # Data rows: left on B, right on last_col, section dividers
    for i in range(len(data_rows)):
        r = first_data_row + i
        for c in range(2, last_col + 1):
            cell = ws.cell(row=r, column=c)
            left = THIN_SIDE if c == 2 else NO_SIDE
            right = THIN_SIDE if c == last_col else (HAIR_SIDE if c in divider_cols else NO_SIDE)
            cell.border = Border(left=left, right=right)

    # Row before totals: bottom thin border
    pre_total_row = total_row - 1
    for c in range(2, last_col + 1):
        cell = ws.cell(row=pre_total_row, column=c)
        existing = cell.border
        cell.border = Border(
            left=existing.left, right=existing.right,
            top=existing.top, bottom=THIN_SIDE
        )

    # Total row: bottom thin border, bold
    for c in range(2, last_col + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.font = BOLD_FONT
        left = THIN_SIDE if c == 2 else NO_SIDE
        right = THIN_SIDE if c == last_col else (HAIR_SIDE if c in divider_cols else NO_SIDE)
        cell.border = Border(left=left, right=right, bottom=THIN_SIDE)

    # Column widths
    ws.column_dimensions["A"].width = 2  # Spacer
    ws.column_dimensions["B"].width = 16  # Unit Type
    for sdef in section_defs:
        for i in range(sdef["ncols"]):
            col_letter = get_column_letter(sdef["start_col"] + i)
            ws.column_dimensions[col_letter].width = 11


def write_checks_sheet(wb: openpyxl.Workbook, agg_data: dict, summary_total_row: int,
                       std_last_row: int, dedup_col: str = "Q",
                       l2l_total_row: int | None = None,
                       le_total_row: int | None = None,
                       occ_total_row: int | None = None):
    """Write the 'Checks' sheet with reconciliation formulas."""
    ws = wb.create_sheet("Checks")

    ws.cell(row=1, column=1, value="Reconciliation Checks").font = BOLD_FONT
    ws.cell(row=1, column=1).alignment = LEFT_ALIGN

    headers = ["Check", "Expected", "Actual", "Pass"]
    for i, h in enumerate(headers):
        ws.cell(row=3, column=i + 1, value=h).font = BOLD_FONT

    dc = dedup_col  # shorthand
    occ_range = f'Standardized!$H$2:$H${std_last_row}'
    dedup_range = f'Standardized!${dc}$2:${dc}${std_last_row}'

    r = 4  # current row tracker

    # Check 1: Total units
    ws.cell(row=r, column=1, value="Total Units (Summary vs Standardized)")
    ws.cell(row=r, column=2, value=f"=Summary!C{summary_total_row}")
    ws.cell(row=r, column=3, value=f'=COUNTIFS({dedup_range},"")')
    ws.cell(row=r, column=4, value=f"=B{r}=C{r}")
    r += 1

    # Check 2: Total SF
    ws.cell(row=r, column=1, value="Total SF (Summary vs Standardized)")
    ws.cell(row=r, column=2, value=f"=Summary!E{summary_total_row}")
    ws.cell(row=r, column=3, value=f'=SUMIFS(Standardized!$F$2:$F${std_last_row},{dedup_range},"")')
    ws.cell(row=r, column=4, value=f"=B{r}=C{r}")
    r += 1

    # Check 3: Total occupied lease rent
    ws.cell(row=r, column=1, value="Total Occupied Lease Rent (Summary vs Standardized)")
    ws.cell(row=r, column=2, value=f"=Summary!K{summary_total_row}*Summary!G{summary_total_row}")
    ws.cell(row=r, column=3, value=f'=SUMIFS(Standardized!$O$2:$O${std_last_row},{occ_range},"Occupied",{dedup_range},"")')
    ws.cell(row=r, column=4, value=f"=ROUND(B{r},0)=ROUND(C{r},0)")
    r += 1

    # Check 4: Occupied units (L2L vs Standardized)
    if l2l_total_row is not None:
        ws.cell(row=r, column=1, value="Occupied Units (L2L vs Standardized)")
        ws.cell(row=r, column=2, value=f"='Loss-to-Lease'!B{l2l_total_row}")
        ws.cell(row=r, column=3, value=f'=COUNTIFS({occ_range},"Occupied",{dedup_range},"")')
        ws.cell(row=r, column=4, value=f"=B{r}=C{r}")
        r += 1

    # Check 5: Occupancy status breakdown (Occupied + Vacant = Total)
    if occ_total_row is not None:
        ws.cell(row=r, column=1, value="Occupancy Breakdown (Occ + Vac = Total)")
        ws.cell(row=r, column=2, value=f"='Occupancy'!B{occ_total_row}")
        ws.cell(row=r, column=3, value=f"='Occupancy'!C{occ_total_row}+'Occupancy'!D{occ_total_row}")
        ws.cell(row=r, column=4, value=f"=B{r}=C{r}")
        r += 1

    # Check 6: Lease Expirations total occupied
    if le_total_row is not None:
        ws.cell(row=r, column=1, value="Lease Exp Total Occupied (Std)")
        ws.cell(row=r, column=2, value=f'=COUNTIFS({occ_range},"Occupied",{dedup_range},"")')
        ws.cell(row=r, column=3, value=f"='Lease Expirations'!B{le_total_row}")
        ws.cell(row=r, column=4).value = f'="Exp=" & C{r} & " / Occ=" & B{r}'
        r += 1

    # Dedup section
    r += 1
    ws.cell(row=r, column=1, value="Deduped Units Count").font = BOLD_FONT
    ws.cell(row=r, column=2, value=f'=COUNTIF({dedup_range},"DROPPED")')
    r += 2

    ws.cell(row=r, column=1, value="Deduped Unit Details").font = BOLD_FONT
    r += 1
    dedup_headers = ["Unit", "Dropped Status", "Dropped Row", "Kept Status", "Kept Row"]
    for i, h in enumerate(dedup_headers):
        ws.cell(row=r, column=i + 1, value=h).font = BOLD_FONT
    r += 1

    dedup_report = agg_data.get("dedup_report", [])
    for j, entry in enumerate(dedup_report):
        ws.cell(row=r + j, column=1, value=entry["unit"])
        ws.cell(row=r + j, column=2, value=entry["dropped_status"])
        ws.cell(row=r + j, column=3, value=entry["dropped_row"])
        ws.cell(row=r + j, column=4, value=entry["kept_status"])
        ws.cell(row=r + j, column=5, value=entry["kept_row"])

    # Column widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12

    return ws


def write_loss_to_lease_sheet(wb: openpyxl.Workbook, agg_data: dict, std_last_row: int,
                              dedup_col: str = "Q"):
    """Write the 'Loss-to-Lease' sheet: per-unit-type market vs in-place rent analysis."""
    ws = wb.create_sheet("Loss-to-Lease")
    unit_type_order = agg_data["unit_type_order"]
    dc = dedup_col

    # Title
    ws.cell(row=1, column=1, value="Loss-to-Lease Analysis").font = BOLD_FONT

    # Headers
    headers = ["Unit Type", "Units", "Avg Market Rent", "Avg In-Place Rent",
               "Loss/Gain per Unit", "Loss/Gain %", "Total Annual Loss/Gain"]
    for i, h in enumerate(headers):
        ws.cell(row=3, column=i + 1, value=h).font = HEADER_FONT

    # Criteria building helpers
    def type_crit(t):
        return f'Standardized!$D$2:$D${std_last_row},"{t}"'

    def base_crit(t):
        return f'{type_crit(t)},Standardized!${dc}$2:${dc}${std_last_row},""'

    occ_crit = f'Standardized!$H$2:$H${std_last_row},"Occupied"'
    market_range = f"Standardized!$N$2:$N${std_last_row}"
    lease_range = f"Standardized!$O$2:$O${std_last_row}"

    for i, utype in enumerate(unit_type_order):
        r = 4 + i
        bc = base_crit(utype)
        occ_bc = f"{bc},{occ_crit}"

        ws.cell(row=r, column=1, value=utype).font = DATA_FONT
        # Units (occupied)
        ws.cell(row=r, column=2, value=f"=COUNTIFS({occ_bc})")
        # Avg Market Rent (all units of type)
        ws.cell(row=r, column=3, value=f"=IFERROR(AVERAGEIFS({market_range},{bc}),0)")
        ws.cell(row=r, column=3).number_format = FMT_CURRENCY
        # Avg In-Place Rent (occupied)
        ws.cell(row=r, column=4, value=f"=IFERROR(AVERAGEIFS({lease_range},{occ_bc}),0)")
        ws.cell(row=r, column=4).number_format = FMT_CURRENCY
        # Loss/Gain per Unit = Avg In-Place - Avg Market
        ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
        ws.cell(row=r, column=5).number_format = FMT_CURRENCY
        # Loss/Gain % = (In-Place - Market) / Market
        ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},0)")
        ws.cell(row=r, column=6).number_format = FMT_PCT
        # Total Annual Loss = Loss per unit * Units * 12
        ws.cell(row=r, column=7, value=f"=E{r}*B{r}*12")
        ws.cell(row=r, column=7).number_format = FMT_CURRENCY

    # Totals row
    total_r = 4 + len(unit_type_order) + 1  # +1 for blank separator
    ws.cell(row=total_r, column=1, value="Total / WA").font = BOLD_FONT

    dedup_crit_all = f'Standardized!${dc}$2:${dc}${std_last_row},""'
    occ_all = f'{occ_crit},{dedup_crit_all}'

    # Total occupied units
    ws.cell(row=total_r, column=2, value=f"=COUNTIFS({occ_all})").font = BOLD_FONT
    # WA Market Rent
    ws.cell(row=total_r, column=3,
            value=f"=IFERROR(SUMIFS({market_range},{dedup_crit_all})/COUNTIFS({dedup_crit_all}),0)").font = BOLD_FONT
    ws.cell(row=total_r, column=3).number_format = FMT_CURRENCY
    # WA In-Place Rent
    ws.cell(row=total_r, column=4,
            value=f"=IFERROR(SUMIFS({lease_range},{occ_all})/COUNTIFS({occ_all}),0)").font = BOLD_FONT
    ws.cell(row=total_r, column=4).number_format = FMT_CURRENCY
    # Loss per unit
    ws.cell(row=total_r, column=5, value=f"=D{total_r}-C{total_r}").font = BOLD_FONT
    ws.cell(row=total_r, column=5).number_format = FMT_CURRENCY
    # Loss %
    ws.cell(row=total_r, column=6, value=f"=IFERROR(E{total_r}/C{total_r},0)").font = BOLD_FONT
    ws.cell(row=total_r, column=6).number_format = FMT_PCT
    # Total Annual
    data_rows = list(range(4, 4 + len(unit_type_order)))
    annual_refs = "+".join(f"G{r}" for r in data_rows)
    ws.cell(row=total_r, column=7, value=f"={annual_refs}").font = BOLD_FONT
    ws.cell(row=total_r, column=7).number_format = FMT_CURRENCY

    # Mark-to-Market summary
    mtm_r = total_r + 2
    ws.cell(row=mtm_r, column=1, value="Mark-to-Market Summary").font = BOLD_FONT
    ws.cell(row=mtm_r + 1, column=1, value="Current Total Monthly Rent (Occupied)")
    ws.cell(row=mtm_r + 1, column=2, value=f"=SUMIFS({lease_range},{occ_all})")
    ws.cell(row=mtm_r + 1, column=2).number_format = FMT_CURRENCY
    ws.cell(row=mtm_r + 2, column=1, value="Potential at Market (All Units)")
    ws.cell(row=mtm_r + 2, column=2, value=f"=SUMIFS({market_range},{dedup_crit_all})")
    ws.cell(row=mtm_r + 2, column=2).number_format = FMT_CURRENCY
    ws.cell(row=mtm_r + 3, column=1, value="Monthly Loss-to-Lease")
    ws.cell(row=mtm_r + 3, column=2, value=f"=B{mtm_r + 1}-B{mtm_r + 2}")
    ws.cell(row=mtm_r + 3, column=2).number_format = FMT_CURRENCY
    ws.cell(row=mtm_r + 4, column=1, value="Annual Loss-to-Lease")
    ws.cell(row=mtm_r + 4, column=2, value=f"=B{mtm_r + 3}*12")
    ws.cell(row=mtm_r + 4, column=2).number_format = FMT_CURRENCY

    # Checks section
    chk_r = mtm_r + 6
    ws.cell(row=chk_r, column=1, value="Checks").font = BOLD_FONT
    ws.cell(row=chk_r + 1, column=1, value="Occupied Units (L2L)")
    ws.cell(row=chk_r + 1, column=2, value=f"=B{total_r}")
    ws.cell(row=chk_r + 2, column=1, value="Occupied Units (Standardized)")
    ws.cell(row=chk_r + 2, column=2, value=f'=COUNTIFS({occ_all})')
    ws.cell(row=chk_r + 3, column=1, value="Match")
    ws.cell(row=chk_r + 3, column=2, value=f"=B{chk_r + 1}=B{chk_r + 2}")
    # Per-type unit sum check
    ws.cell(row=chk_r + 4, column=1, value="Sum of Type Units")
    unit_refs = "+".join(f"B{r}" for r in data_rows)
    ws.cell(row=chk_r + 4, column=2, value=f"={unit_refs}")
    ws.cell(row=chk_r + 5, column=1, value="Sum = Total Match")
    ws.cell(row=chk_r + 5, column=2, value=f"=B{chk_r + 4}=B{total_r}")

    # Column widths
    ws.column_dimensions["A"].width = 34
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 18

    return total_r


def write_other_income_sheet(wb: openpyxl.Workbook, agg_data: dict, std_last_row: int,
                             charge_std_map: dict, dedup_col: str = "Q"):
    """Write the 'Other Income' sheet: per-unit-type avg and total for each charge column."""
    ws = wb.create_sheet("Other Income")
    unit_type_order = agg_data["unit_type_order"]
    dc = dedup_col

    if not charge_std_map:
        ws.cell(row=1, column=1, value="No charge columns detected in rent roll.").font = DATA_FONT
        return ws

    charge_names = list(charge_std_map.keys())

    # Title
    ws.cell(row=1, column=1, value="Other Income Analysis").font = BOLD_FONT

    # Headers: Unit Type | then for each charge: # <name> | Avg <name> | Total <name>
    ws.cell(row=3, column=1, value="Unit Type").font = HEADER_FONT
    col = 2
    charge_col_positions = {}  # charge_name -> (count_col, avg_col, total_col)
    for name in charge_names:
        ws.cell(row=3, column=col, value=f"# {name}").font = HEADER_FONT
        ws.cell(row=3, column=col + 1, value=f"Avg {name}").font = HEADER_FONT
        ws.cell(row=3, column=col + 2, value=f"Total {name}").font = HEADER_FONT
        charge_col_positions[name] = (col, col + 1, col + 2)
        col += 3

    # Data rows
    for i, utype in enumerate(unit_type_order):
        r = 4 + i
        type_crit = f'Standardized!$D$2:$D${std_last_row},"{utype}"'
        base_crit = f'{type_crit},Standardized!${dc}$2:${dc}${std_last_row},""'

        ws.cell(row=r, column=1, value=utype).font = DATA_FONT

        for name in charge_names:
            std_letter = charge_std_map[name]
            charge_range = f"Standardized!${std_letter}$2:${std_letter}${std_last_row}"
            count_col, avg_col, total_col = charge_col_positions[name]

            # # Paying (non-zero, non-empty)
            ws.cell(row=r, column=count_col,
                    value=f'=COUNTIFS({charge_range},"<>0",{charge_range},"<>",{base_crit})')
            ws.cell(row=r, column=avg_col,
                    value=f"=IFERROR(AVERAGEIFS({charge_range},{base_crit}),0)")
            ws.cell(row=r, column=avg_col).number_format = FMT_CURRENCY
            ws.cell(row=r, column=total_col,
                    value=f"=SUMIFS({charge_range},{base_crit})")
            ws.cell(row=r, column=total_col).number_format = FMT_CURRENCY

    # Total row
    total_r = 4 + len(unit_type_order) + 1
    ws.cell(row=total_r, column=1, value="Total / WA").font = BOLD_FONT
    dedup_crit_all = f'Standardized!${dc}$2:${dc}${std_last_row},""'

    for name in charge_names:
        std_letter = charge_std_map[name]
        charge_range = f"Standardized!${std_letter}$2:${std_letter}${std_last_row}"
        count_col, avg_col, total_col = charge_col_positions[name]

        ws.cell(row=total_r, column=count_col,
                value=f'=COUNTIFS({charge_range},"<>0",{charge_range},"<>",{dedup_crit_all})').font = BOLD_FONT
        ws.cell(row=total_r, column=avg_col,
                value=f"=IFERROR(SUMIFS({charge_range},{dedup_crit_all})/COUNTIFS({dedup_crit_all}),0)").font = BOLD_FONT
        ws.cell(row=total_r, column=avg_col).number_format = FMT_CURRENCY
        ws.cell(row=total_r, column=total_col,
                value=f"=SUMIFS({charge_range},{dedup_crit_all})").font = BOLD_FONT
        ws.cell(row=total_r, column=total_col).number_format = FMT_CURRENCY

    # Checks section
    chk_r = total_r + 2
    ws.cell(row=chk_r, column=1, value="Checks").font = BOLD_FONT
    ws.cell(row=chk_r + 1, column=1, value="Total Units")
    ws.cell(row=chk_r + 1, column=2, value=f'=COUNTIFS({dedup_crit_all})')
    for name in charge_names:
        count_col, _, total_col = charge_col_positions[name]
        total_letter = get_column_letter(total_col)
        ws.cell(row=chk_r + 2, column=1, value=f"Sum check {name}")
        # Sum of per-type totals should match the grand total
        data_row_refs = "+".join(f"{total_letter}{4 + i}" for i in range(len(unit_type_order)))
        ws.cell(row=chk_r + 2 + list(charge_names).index(name), column=1,
                value=f"Sum check: {name}")
        ws.cell(row=chk_r + 2 + list(charge_names).index(name), column=2,
                value=f"={data_row_refs}")
        ws.cell(row=chk_r + 2 + list(charge_names).index(name), column=2).number_format = FMT_CURRENCY
        ws.cell(row=chk_r + 2 + list(charge_names).index(name), column=3,
                value=f"={total_letter}{total_r}")
        ws.cell(row=chk_r + 2 + list(charge_names).index(name), column=3).number_format = FMT_CURRENCY
        ws.cell(row=chk_r + 2 + list(charge_names).index(name), column=4,
                value=f"=ROUND(B{chk_r + 2 + list(charge_names).index(name)},0)=ROUND(C{chk_r + 2 + list(charge_names).index(name)},0)")

    # Column widths
    ws.column_dimensions["A"].width = 20
    for c_idx in range(2, col):
        ws.column_dimensions[get_column_letter(c_idx)].width = 14

    return ws


def write_lease_expirations_sheet(wb: openpyxl.Workbook, agg_data: dict, std_last_row: int,
                                  as_of_date, dedup_col: str = "Q"):
    """Write the 'Lease Expirations' sheet: monthly buckets for 18 months from as-of date."""
    from datetime import timedelta
    import calendar

    ws = wb.create_sheet("Lease Expirations")
    dc = dedup_col

    if as_of_date is None:
        ws.cell(row=1, column=1, value="No As-of Date detected — cannot compute lease expirations.").font = DATA_FONT
        return ws

    # Title
    ws.cell(row=1, column=1, value="Lease Expiration Schedule").font = BOLD_FONT

    # Headers
    headers = ["Month", "Expiring Leases", "Total Expiring Rent", "Avg Expiring Rent"]
    for i, h in enumerate(headers):
        ws.cell(row=3, column=i + 1, value=h).font = HEADER_FONT

    # Build 18 monthly buckets starting from the as-of date's month
    lease_end_range = f"Standardized!$M$2:$M${std_last_row}"
    rent_range = f"Standardized!$O$2:$O${std_last_row}"
    dedup_crit = f'Standardized!${dc}$2:${dc}${std_last_row},""'
    occ_crit = f'Standardized!$H$2:$H${std_last_row},"Occupied"'

    year = as_of_date.year
    month = as_of_date.month

    for i in range(18):
        r = 4 + i
        # Current bucket month/year
        m = ((month - 1 + i) % 12) + 1
        y = year + (month - 1 + i) // 12
        last_day = calendar.monthrange(y, m)[1]

        label = f"{calendar.month_abbr[m]} {y}"
        ws.cell(row=r, column=1, value=label).font = DATA_FONT

        # Lease End >= first of month AND <= last of month
        gte_crit = f'{lease_end_range},">=" & DATE({y},{m},1)'
        lte_crit = f'{lease_end_range},"<=" & DATE({y},{m},{last_day})'

        all_crit = f"{dedup_crit},{occ_crit},{gte_crit},{lte_crit}"

        # Expiring Leases = COUNTIFS
        ws.cell(row=r, column=2, value=f"=COUNTIFS({all_crit})")
        # Total Expiring Rent = SUMIFS
        ws.cell(row=r, column=3, value=f"=SUMIFS({rent_range},{all_crit})")
        ws.cell(row=r, column=3).number_format = FMT_CURRENCY
        # Avg Expiring Rent = AVERAGEIFS
        ws.cell(row=r, column=4, value=f"=IFERROR(AVERAGEIFS({rent_range},{all_crit}),0)")
        ws.cell(row=r, column=4).number_format = FMT_CURRENCY

    # Total row
    total_r = 4 + 18 + 1
    ws.cell(row=total_r, column=1, value="Total").font = BOLD_FONT
    data_rows = list(range(4, 4 + 18))
    ws.cell(row=total_r, column=2, value="=" + "+".join(f"B{r}" for r in data_rows)).font = BOLD_FONT
    ws.cell(row=total_r, column=3, value="=" + "+".join(f"C{r}" for r in data_rows)).font = BOLD_FONT
    ws.cell(row=total_r, column=3).number_format = FMT_CURRENCY
    ws.cell(row=total_r, column=4, value=f"=IFERROR(C{total_r}/B{total_r},0)").font = BOLD_FONT
    ws.cell(row=total_r, column=4).number_format = FMT_CURRENCY

    # Checks section
    chk_r = total_r + 2
    ws.cell(row=chk_r, column=1, value="Checks").font = BOLD_FONT
    occ_range = f"Standardized!$H$2:$H${std_last_row}"
    ws.cell(row=chk_r + 1, column=1, value="Total Occupied")
    ws.cell(row=chk_r + 1, column=2,
            value=f'=COUNTIFS({occ_range},"Occupied",{dedup_crit})')
    ws.cell(row=chk_r + 2, column=1, value="Expiring in 18 Mo")
    ws.cell(row=chk_r + 2, column=2, value=f"=B{total_r}")
    ws.cell(row=chk_r + 3, column=1, value="Beyond 18 Mo / MTM")
    ws.cell(row=chk_r + 3, column=2, value=f"=B{chk_r + 1}-B{chk_r + 2}")
    ws.cell(row=chk_r + 4, column=1, value="Sanity (Occ = Exp + Beyond)")
    ws.cell(row=chk_r + 4, column=2, value=f"=B{chk_r + 1}=(B{chk_r + 2}+B{chk_r + 3})")

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18

    return total_r


def write_occupancy_sheet(wb: openpyxl.Workbook, agg_data: dict, std_last_row: int,
                          charge_std_map: dict, dedup_col: str = "Q"):
    """Write the 'Occupancy' sheet: per-unit-type status breakdown and economic occupancy."""
    ws = wb.create_sheet("Occupancy")
    unit_type_order = agg_data["unit_type_order"]
    dc = dedup_col

    # Title
    ws.cell(row=1, column=1, value="Occupancy Analysis").font = BOLD_FONT

    # Headers
    headers = ["Unit Type", "Total", "Occupied", "Vacant", "Applicant",
               "NTV", "Pending Renewal", "Physical Occ %", "Economic Occ %",
               "% Leased", "% Leased ex. NTVs"]
    for i, h in enumerate(headers):
        ws.cell(row=3, column=i + 1, value=h).font = HEADER_FONT

    status_range = f"Standardized!$G$2:$G${std_last_row}"
    market_range = f"Standardized!$N$2:$N${std_last_row}"
    lease_range = f"Standardized!$O$2:$O${std_last_row}"
    occ_range = f"Standardized!$H$2:$H${std_last_row}"

    for i, utype in enumerate(unit_type_order):
        r = 4 + i
        type_crit = f'Standardized!$D$2:$D${std_last_row},"{utype}"'
        base_crit = f'{type_crit},Standardized!${dc}$2:${dc}${std_last_row},""'

        ws.cell(row=r, column=1, value=utype).font = DATA_FONT
        # Total
        ws.cell(row=r, column=2, value=f"=COUNTIFS({base_crit})")
        # Occupied (Occupancy = "Occupied")
        ws.cell(row=r, column=3, value=f'=COUNTIFS({base_crit},{occ_range},"Occupied")')
        # Vacant
        ws.cell(row=r, column=4, value=f'=COUNTIFS({base_crit},{status_range},"Vacant")')
        # Applicant
        ws.cell(row=r, column=5, value=f'=COUNTIFS({base_crit},{status_range},"Applicant")')
        # NTV (Occupied-NTV)
        ws.cell(row=r, column=6, value=f'=COUNTIFS({base_crit},{status_range},"Occupied-NTV")')
        # Pending Renewal
        ws.cell(row=r, column=7, value=f'=COUNTIFS({base_crit},{status_range},"Pending Renewal")')
        # Physical Occ % = Occupied / Total
        ws.cell(row=r, column=8, value=f"=IFERROR(C{r}/B{r},0)")
        ws.cell(row=r, column=8).number_format = FMT_PCT
        # Economic Occ % = total in-place rent / total market rent
        occ_crit = f'{occ_range},"Occupied"'
        ws.cell(row=r, column=9,
                value=f"=IFERROR(SUMIFS({lease_range},{base_crit},{occ_crit})/SUMIFS({market_range},{base_crit}),0)")
        ws.cell(row=r, column=9).number_format = FMT_PCT
        # % Leased = Occupied / Total (includes NTV, Applicant, Pending Renewal)
        ws.cell(row=r, column=10, value=f"=IFERROR(C{r}/B{r},0)")
        ws.cell(row=r, column=10).number_format = FMT_PCT
        # % Leased ex. NTVs = (Occupied - NTV) / Total
        ws.cell(row=r, column=11, value=f"=IFERROR((C{r}-F{r})/B{r},0)")
        ws.cell(row=r, column=11).number_format = FMT_PCT

    # Total row
    total_r = 4 + len(unit_type_order) + 1
    ws.cell(row=total_r, column=1, value="Total").font = BOLD_FONT
    data_rows = list(range(4, 4 + len(unit_type_order)))
    for c in range(2, 8):  # Columns B through G: SUM
        refs = "+".join(f"{get_column_letter(c)}{r}" for r in data_rows)
        ws.cell(row=total_r, column=c, value=f"={refs}").font = BOLD_FONT
    # Physical Occ %
    ws.cell(row=total_r, column=8, value=f"=IFERROR(C{total_r}/B{total_r},0)").font = BOLD_FONT
    ws.cell(row=total_r, column=8).number_format = FMT_PCT
    # Economic Occ %
    dedup_crit_all = f'Standardized!${dc}$2:${dc}${std_last_row},""'
    occ_crit_all = f'{occ_range},"Occupied"'
    ws.cell(row=total_r, column=9,
            value=f"=IFERROR(SUMIFS({lease_range},{occ_crit_all},{dedup_crit_all})/SUMIFS({market_range},{dedup_crit_all}),0)").font = BOLD_FONT
    ws.cell(row=total_r, column=9).number_format = FMT_PCT
    # % Leased
    ws.cell(row=total_r, column=10, value=f"=IFERROR(C{total_r}/B{total_r},0)").font = BOLD_FONT
    ws.cell(row=total_r, column=10).number_format = FMT_PCT
    # % Leased ex. NTVs
    ws.cell(row=total_r, column=11, value=f"=IFERROR((C{total_r}-F{total_r})/B{total_r},0)").font = BOLD_FONT
    ws.cell(row=total_r, column=11).number_format = FMT_PCT

    # Concessions section
    conc_r = total_r + 3
    ws.cell(row=conc_r, column=1, value="Concessions Summary").font = BOLD_FONT

    # Try to find EMPLCRED in charge columns
    emplcred_col = charge_std_map.get("EMPLCRED")
    if emplcred_col:
        emplcred_range = f"Standardized!${emplcred_col}$2:${emplcred_col}${std_last_row}"
        ws.cell(row=conc_r + 1, column=1, value="Units with Employee Credit")
        ws.cell(row=conc_r + 1, column=2,
                value=f'=COUNTIFS({dedup_crit_all},{emplcred_range},"<>0",{emplcred_range},"<>")')
        ws.cell(row=conc_r + 2, column=1, value="Total Employee Credit Amount")
        ws.cell(row=conc_r + 2, column=2,
                value=f"=SUMIFS({emplcred_range},{dedup_crit_all})")
        ws.cell(row=conc_r + 2, column=2).number_format = FMT_CURRENCY
    else:
        ws.cell(row=conc_r + 1, column=1, value="No EMPLCRED column detected in rent roll.").font = DATA_FONT

    # Checks section
    chk_r = conc_r + 4
    ws.cell(row=chk_r, column=1, value="Checks").font = BOLD_FONT
    ws.cell(row=chk_r + 1, column=1, value="Total Units (Occupancy)")
    ws.cell(row=chk_r + 1, column=2, value=f"=B{total_r}")
    ws.cell(row=chk_r + 2, column=1, value="Occupied + Vacant")
    ws.cell(row=chk_r + 2, column=2, value=f"=C{total_r}+D{total_r}")
    ws.cell(row=chk_r + 3, column=1, value="Match")
    ws.cell(row=chk_r + 3, column=2, value=f"=B{chk_r + 1}=B{chk_r + 2}")

    # Column widths
    ws.column_dimensions["A"].width = 22
    for c in "BCDEFGHIJK":
        ws.column_dimensions[c].width = 16

    return total_r


def generate_output(agg_data: dict, column_map: dict, raw_wb_bytes: bytes,
                    sheet_name: str, date_ranges: list[dict],
                    as_of_date=None) -> bytes:
    """Generate the full output workbook and return as bytes."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Raw Data
    write_raw_data_sheet(wb, raw_wb_bytes, sheet_name)

    # Sheet 2: Standardized (returns layout info for downstream sheets)
    _, std_columns, charge_std_map = write_standardized_sheet(wb, agg_data, column_map)

    # Determine dedup column letter from the built layout
    dedup_col = "Q"  # default
    for letter, header in std_columns:
        if header == "Dedup Flag":
            dedup_col = letter
            break

    # Sheet 3: Summary
    unit_type_order = agg_data["unit_type_order"]
    non_reno = [t for t in unit_type_order if not t.endswith(" Reno")]
    reno = [t for t in unit_type_order if t.endswith(" Reno")]
    # Data rows = non_reno + separator + reno + separator
    n_data_rows = len(non_reno) + 1 + len(reno) + 1
    total_row = 6 + n_data_rows
    std_last_row = len(agg_data["df"]) + 1  # +1 for header

    write_summary_sheet(wb, agg_data, date_ranges, dedup_col)

    # Sheet 4: Loss-to-Lease
    l2l_total_row = write_loss_to_lease_sheet(wb, agg_data, std_last_row, dedup_col)

    # Sheet 5: Other Income
    write_other_income_sheet(wb, agg_data, std_last_row, charge_std_map, dedup_col)

    # Sheet 6: Lease Expirations
    le_total_row = write_lease_expirations_sheet(wb, agg_data, std_last_row, as_of_date, dedup_col)

    # Sheet 7: Occupancy
    occ_total_row = write_occupancy_sheet(wb, agg_data, std_last_row, charge_std_map, dedup_col)

    # Sheet 8: Checks
    write_checks_sheet(wb, agg_data, total_row, std_last_row, dedup_col,
                       l2l_total_row=l2l_total_row,
                       le_total_row=le_total_row if isinstance(le_total_row, int) else None,
                       occ_total_row=occ_total_row if isinstance(occ_total_row, int) else None)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
