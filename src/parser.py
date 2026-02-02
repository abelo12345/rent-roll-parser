"""Step 1: LLM-powered column identification & data extraction from rent roll Excel files."""

import json
import re
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
import pandas as pd

# Canonical status values
VALID_STATUSES = {"Occupied", "Vacant", "Applicant", "Occupied-NTV", "Pending Renewal", "Model"}

# Minimal hardcoded fallback for the most universal status values.
# The AI-provided status_mapping (from identify_columns) handles all format-specific
# abbreviations and PMS codes dynamically.
_BASIC_STATUS_ALIASES = {
    "occupied": "Occupied",
    "vacant": "Vacant",
    "applicant": "Applicant",
    "model": "Model",
}


_AS_OF_PATTERNS = [
    r'[Aa]s\s+[Oo]f\s+[Dd]ate[:\s]+(\d{1,2}/\d{1,2}/\d{2,4})',  # "As of Date: MM/DD/YYYY"
    r'[Aa]s\s+[Oo]f\s*=\s*(\d{1,2}/\d{1,2}/\d{2,4})',            # Yardi "As Of = MM/DD/YYYY"
    r'[Aa]s\s+[Oo]f\s*:\s*(\d{1,2}/\d{1,2}/\d{2,4})',            # generic colon
    r'[Aa]s\s+[Oo]f\s+(\d{1,2}/\d{1,2}/\d{2,4})',                # no separator
]


def extract_as_of_date(ws) -> datetime | None:
    """Scan the first 10 rows for an as-of date pattern and return it."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for val in row:
            if val is None:
                continue
            text = str(val)
            for pattern in _AS_OF_PATTERNS:
                m = re.search(pattern, text)
                if m:
                    return normalize_date(m.group(1))
    return None


def _pick_data_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    """Pick the sheet with the most data rows."""
    best, best_count = wb.active, 0
    for name in wb.sheetnames:
        ws = wb[name]
        count = sum(1 for row in ws.iter_rows(min_row=1, max_row=ws.max_row) if any(c.value is not None for c in row))
        if count > best_count:
            best, best_count = ws, count
    return best


def _col_letter(idx: int) -> str:
    """1-based index to Excel column letter."""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _col_index(letter: str) -> int:
    """Excel column letter to 1-based index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - 64)
    return result


def _rows_to_text(ws, max_rows: int = 50) -> str:
    """Convert first N rows of a worksheet to a text table for the LLM."""
    lines = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=False), start=1):
        parts = []
        for cell in row:
            col_letter = _col_letter(cell.column)
            val = cell.value
            if val is None:
                continue
            parts.append(f"{col_letter}{row_idx}={repr(val)}")
        if parts:
            lines.append(f"Row {row_idx}: " + " | ".join(parts))
    return "\n".join(lines)


def _rows_to_text_range(ws, start_row: int, end_row: int) -> str:
    """Convert a range of rows to text for the LLM."""
    lines = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=start_row, max_row=end_row, values_only=False),
        start=start_row,
    ):
        parts = []
        for cell in row:
            col_letter = _col_letter(cell.column)
            val = cell.value
            if val is None:
                continue
            parts.append(f"{col_letter}{row_idx}={repr(val)}")
        if parts:
            lines.append(f"Row {row_idx}: " + " | ".join(parts))
    return "\n".join(lines)


def identify_columns(client, ws) -> dict:
    """Use Claude to identify the column mapping and data boundaries in the rent roll."""
    preview = _rows_to_text(ws, max_rows=50)

    # Also include the last 20 rows so the AI can detect footer/summary patterns
    tail_start = max(51, ws.max_row - 19)
    tail_preview = _rows_to_text_range(ws, tail_start, ws.max_row)

    prompt = f"""You are analyzing a multifamily rent roll spreadsheet. Below are the first 50 rows with cell references and values.

{preview}

--- LAST {ws.max_row - tail_start + 1} ROWS (rows {tail_start}-{ws.max_row}) ---

{tail_preview}

Identify the following and return ONLY valid JSON (no markdown fences, no explanation):
{{
  "header_row": <row number containing column headers>,
  "columns": {{
    "unit": "<column letter for unit number/ID>",
    "floorplan": "<column letter for floorplan code>",
    "sqft": "<column letter for square footage>",
    "status": "<column letter for unit/lease status>",
    "tenant_name": "<column letter for tenant name>",
    "move_in": "<column letter for move-in date>",
    "move_out": "<column letter for move-out date>",
    "lease_start": "<column letter for lease start date>",
    "lease_end": "<column letter for lease end date>",
    "market_rent": "<column letter for market rent>",
    "lease_rent": "<column letter for lease/contract rent>",
    "total_billing": "<column letter for total billing/charges>"
  }},
  "charge_columns": {{
    "<CHARGE_NAME>": "<column letter>",
    ...
  }},
  "data_start_row": <first row of unit data>,
  "format": "<single_row or multi_row>",
  "status_column_exists": <true or false>,
  "multi_row_config": <null or object>,
  "section_dividers": <[] or list of objects>,
  "status_mapping": {{
    "<raw_status_value>": "<canonical_status>",
    ...
  }},
  "footer_patterns": ["<lowercase text pattern>", ...],
  "section_keywords": {{
    "applicants": ["<keyword>", ...],
    "current": ["<keyword>", ...]
  }}
}}

FORMAT DETECTION:
- "single_row": Each unit occupies exactly one row. All fields (unit, tenant, rent, charges) are in the same row.
- "multi_row": Each unit spans multiple rows. Typically the first row has the unit info (unit number, tenant, dates, sqft, etc.) and subsequent rows contain individual charge detail lines (one charge per row with a code column and an amount column), followed by a "Total" row that sums the charges. There is usually a blank row separator between units.
  If format is "multi_row", set "multi_row_config" to:
  {{
    "charge_code_column": "<column letter containing charge code/description like 'rent', 'pet rent', etc.>",
    "charge_amount_column": "<column letter containing the charge dollar amount>",
    "total_row_indicator": "<text that marks a total row, e.g. 'Total' or 'Unit Total'>",
    "rent_charge_codes": ["<charge codes that represent base rent, e.g. 'rent', 'Rent'>"]
  }}

STATUS COLUMN:
- "status_column_exists": true if there is an explicit status/lease-status column (with values like "Current", "Occupied", "Vacant", "Notice", etc.)
- "status_column_exists": false if there is no dedicated status column. In this case set columns.status to null.

SECTION DIVIDERS:
- Some rent rolls have section headers that divide units into groups like "Current Residents", "Applicants", "Future Residents", etc.
- If you see any such section divider rows, return them as:
  [{{"row": <row number>, "text": "<section header text>", "section_type": "<current or applicants>"}}]
- "section_type" should be "current" for sections containing current/active residents and "applicants" for sections containing future/applicant/pre-lease residents.
- If there are no section dividers, return an empty list [].

STATUS MAPPING:
- Look at the status/lease-status column values visible in the preview rows.
- Map every unique raw status value to one of these canonical statuses:
  "Occupied", "Vacant", "Applicant", "Occupied-NTV", "Pending Renewal", "Model"
- Include ALL status values you see, even abbreviated PMS codes (e.g., "OC" → "Occupied", "VR" → "Vacant", "NU" → "Occupied-NTV")
- Common PMS abbreviations: OC/Current/Leased = Occupied, VR/VU/VN = Vacant, NR/NU/NN/NTV/Notice = Occupied-NTV, AP/Pre-leased = Applicant, PR/Renewed = Pending Renewal, MO = Model
- Return as: {{"OC": "Occupied", "VR": "Vacant", ...}}
- If no status column exists (status_column_exists is false), return {{}}

FOOTER PATTERNS:
- Identify any text patterns visible in the preview that mark summary, footer, or legend rows
  (e.g., "Summary Groups", "Total Units", "Status Legend", "Grand Total", "Charge Summary")
- Return as a list of lowercase strings that can be used for substring matching
- If none visible, return []

SECTION KEYWORDS:
- If you see section header rows, also return generic keywords to help detect similar sections beyond the 50-row preview
- Format: {{"applicants": ["future residents", ...], "current": ["current residents", ...]}}
- If no section headers are visible, return {{}}

TWO-ROW HEADERS:
- If the header spans two rows (e.g. row 5 and row 6), use the FIRST header row as "header_row" and make sure "data_start_row" points to the first actual data row after both header rows.

Important:
- "unit" is the apartment/unit number (like "A101", "B205", "0101"), not a resident or lease ID
- "lease_rent" is the contract/actual rent the tenant pays, NOT market rent
- "total_billing" is total charges including rent + utilities + fees
- If a column doesn't exist, use null for its value
- Do NOT include a "data_end_row" — I will detect that programmatically
- For "single_row" format: "charge_columns" should list every individual charge/fee column BETWEEN lease_rent and total_billing (e.g. RENT, WATER, DWP, PETRENT, PARKING, AMENITY, EMPLCRED, GUEST, STORAGE, MODEL, MTOM). Use the exact header text as the key. Omit this field if no charge columns exist.
- For "multi_row" format: set "charge_columns" to {{}} (empty) since charges are in rows not columns."""

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2048,
        tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}],
        messages=[{"role": "user", "content": prompt}],
    )
    text = "".join(b.text for b in response.content if hasattr(b, "text")).strip()
    # Strip markdown fences if present
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    # Extract JSON object from surrounding commentary (web search may add extra text)
    match = re.search(r"\{", text)
    if match:
        text = text[match.start():]
        # Find matching closing brace
        depth = 0
        for i, ch in enumerate(text):
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    text = text[: i + 1]
                    break
    return json.loads(text)


def normalize_status(raw: str, ai_status_map: dict[str, str] | None = None) -> str:
    """Normalize a status string using the AI-provided mapping, with minimal fallback."""
    if raw is None:
        return "Vacant"
    cleaned = str(raw).strip()
    lower = cleaned.lower()

    # Try AI-provided mapping first (exact match on lowercased key)
    if ai_status_map:
        if lower in ai_status_map:
            return ai_status_map[lower]
        if cleaned in ai_status_map:
            return ai_status_map[cleaned]

    # Minimal hardcoded fallback
    if lower in _BASIC_STATUS_ALIASES:
        return _BASIC_STATUS_ALIASES[lower]

    # Fallback: return as-is
    return cleaned


def _infer_status(record: dict, section_type: str | None = None) -> str:
    """Infer status when no explicit status column exists."""
    tenant = record.get("tenant_name")
    tenant_str = str(tenant).strip().upper() if tenant else ""

    if tenant_str in ("", "VACANT"):
        return "Vacant"
    if tenant_str == "MODEL":
        return "Model"
    if section_type == "applicants":
        return "Applicant"
    move_out = record.get("move_out")
    if move_out is not None and str(move_out).strip() != "":
        return "Occupied-NTV"
    return "Occupied"


def _excel_serial_to_date(serial) -> datetime | None:
    """Convert Excel serial date number to datetime."""
    if serial is None:
        return None
    try:
        serial = float(serial)
    except (ValueError, TypeError):
        return None
    if serial < 1:
        return None
    # Excel epoch is 1900-01-01, but has the 1900 leap year bug
    base = datetime(1899, 12, 30)
    return base + timedelta(days=serial)


def normalize_date(val) -> datetime | None:
    """Normalize a date value (handles datetime, serial numbers, strings)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        return _excel_serial_to_date(val)
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Try Excel serial from string
    try:
        return _excel_serial_to_date(float(s))
    except (ValueError, TypeError):
        return None


def _normalize_record(record: dict) -> dict:
    """Normalize date and numeric fields in a record in-place, return it."""
    for date_field in ("move_in", "move_out", "lease_start", "lease_end"):
        record[date_field] = normalize_date(record.get(date_field))
    for num_field in ("sqft", "market_rent", "lease_rent", "total_billing"):
        val = record.get(num_field)
        if val is not None:
            try:
                record[num_field] = float(val)
            except (ValueError, TypeError):
                record[num_field] = None
    # Normalize charge_* fields
    for key in list(record.keys()):
        if key.startswith("charge_"):
            val = record[key]
            if val is not None:
                try:
                    record[key] = float(val)
                except (ValueError, TypeError):
                    record[key] = None
    return record


def _get_section_type_for_row(row_idx: int, section_dividers: list[dict]) -> str | None:
    """Determine which section a row belongs to based on section dividers."""
    if not section_dividers:
        return None
    current_section = None
    for div in sorted(section_dividers, key=lambda d: d["row"]):
        if row_idx >= div["row"]:
            current_section = div.get("section_type")
        else:
            break
    return current_section


# Minimal hardcoded footer markers (always checked). The AI-provided footer_patterns
# from identify_columns() supplement these dynamically.
_FOOTER_MARKERS_FALLBACK = {"grand total", "property total"}


def _is_footer_row(ws, row_idx: int, ai_footer_patterns: list[str] | None = None) -> bool:
    """Check if a row matches footer/summary patterns (AI-provided + fallback)."""
    row_text = ""
    for col in range(1, min(ws.max_column + 1, 15)):
        v = ws.cell(row=row_idx, column=col).value
        if v is not None:
            row_text += str(v).lower() + " "

    # Check hardcoded fallback
    for marker in _FOOTER_MARKERS_FALLBACK:
        if marker in row_text:
            return True

    # Check AI-provided patterns
    if ai_footer_patterns:
        for pattern in ai_footer_patterns:
            if pattern.lower() in row_text:
                return True

    return False


def _find_data_end(ws, data_start: int, unit_col_idx: int,
                   ai_footer_patterns: list[str] | None = None) -> int:
    """Scan from data_start to find the last data row (before grand totals/footer)."""
    data_end = data_start
    blank_streak = 0
    for row_idx in range(data_start, ws.max_row + 1):
        if _is_footer_row(ws, row_idx, ai_footer_patterns):
            break

        cell_val = ws.cell(row=row_idx, column=unit_col_idx).value
        if cell_val is not None and str(cell_val).strip() != "":
            data_end = row_idx
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak > 20:
                break
    return data_end


def _compute_section_ranges(
    section_dividers: list[dict], data_start: int, data_end: int
) -> list[tuple[int, int, str | None]]:
    """Compute (start_row, end_row, section_type) ranges from section dividers.

    Each section starts at divider_row + 1 and ends at the next divider - 1
    (or data_end for the last section).  If there are no dividers, returns
    a single range covering data_start to data_end with section_type=None.
    """
    if not section_dividers:
        return [(data_start, data_end, None)]

    sorted_divs = sorted(section_dividers, key=lambda d: d["row"])
    ranges = []
    for i, div in enumerate(sorted_divs):
        start = div["row"] + 1
        if i + 1 < len(sorted_divs):
            end = sorted_divs[i + 1]["row"] - 1
        else:
            end = data_end
        ranges.append((start, end, div.get("section_type")))
    return ranges


# Minimal hardcoded section keywords (fallback). The AI-provided section_keywords
# from identify_columns() supplement these dynamically.
_SECTION_KEYWORDS_FALLBACK = {
    "applicants": ["future residents", "applicant"],
    "current": ["current residents"],
}


def _scan_section_dividers(
    ws, data_start: int, existing_dividers: list[dict],
    ai_section_keywords: dict[str, list[str]] | None = None,
    ai_footer_patterns: list[str] | None = None,
) -> list[dict]:
    """Scan the full sheet for section divider rows the LLM may have missed.

    Uses AI-provided section keywords (with fallback) to find dividers beyond
    the 50-row preview window. Stops scanning at footer rows.
    """
    keywords = ai_section_keywords if ai_section_keywords else _SECTION_KEYWORDS_FALLBACK
    known_rows = {d["row"] for d in existing_dividers}
    dividers = list(existing_dividers)

    for row_idx in range(data_start, ws.max_row + 1):
        # Stop at footer content — summary tables reuse section labels
        if _is_footer_row(ws, row_idx, ai_footer_patterns):
            break

        if row_idx in known_rows:
            continue
        a_val = ws.cell(row=row_idx, column=1).value
        if not a_val or not isinstance(a_val, str):
            continue
        text = a_val.strip()
        lower = text.lower()

        # Check against section keywords (AI-provided or fallback)
        for section_type, kw_list in keywords.items():
            if any(kw in lower for kw in kw_list):
                dividers.append({
                    "row": row_idx,
                    "text": text,
                    "section_type": section_type,
                })
                known_rows.add(row_idx)
                break

    return sorted(dividers, key=lambda d: d["row"])


def _filter_non_unit_rows(records: list[dict]) -> list[dict]:
    """Remove parsed rows that are clearly not real units (e.g., summary/legend rows).

    A row is considered non-unit if it has no sqft, no market_rent, no lease_rent,
    and no tenant_name — i.e., it has no substantive data beyond a unit identifier.
    """
    filtered = []
    for r in records:
        sqft = r.get("sqft")
        market = r.get("market_rent")
        lease = r.get("lease_rent")
        tenant = r.get("tenant_name")
        tenant_str = str(tenant).strip() if tenant else ""

        has_data = (
            (sqft is not None and sqft != 0)
            or (market is not None and market != 0)
            or (lease is not None and lease != 0)
            or (tenant_str != "" and tenant_str.upper() not in ("VACANT", "MODEL", ""))
        )
        if has_data:
            filtered.append(r)
    return filtered


def _resolve_unknown_statuses(client, records: list[dict]) -> list[dict]:
    """If any records have status values not in the canonical set, ask AI to resolve them."""
    unknown_statuses = set()
    for r in records:
        s = r.get("status", "")
        if s and s not in VALID_STATUSES:
            unknown_statuses.add(s)

    if not unknown_statuses:
        return records

    # Make a targeted AI call for just the unknown values
    status_list = ", ".join(f'"{s}"' for s in sorted(unknown_statuses))
    prompt = f"""Map each of these rent roll status values to one of the canonical statuses.

Status values to map: [{status_list}]

Canonical statuses: "Occupied", "Vacant", "Applicant", "Occupied-NTV", "Pending Renewal", "Model"

Return ONLY valid JSON (no markdown fences): {{"<raw>": "<canonical>", ...}}"""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        text = "".join(b.text for b in response.content if hasattr(b, "text")).strip()
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        match = re.search(r"\{", text)
        if match:
            text = text[match.start():]
            depth = 0
            for i, ch in enumerate(text):
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        text = text[: i + 1]
                        break
        text = re.sub(r",\s*([}\]])", r"\1", text)
        fix_map = json.loads(text)

        # Apply the fixes
        for r in records:
            s = r.get("status", "")
            if s in fix_map and fix_map[s] in VALID_STATUSES:
                r["status"] = fix_map[s]
    except Exception:
        pass  # If the resolution call fails, leave statuses as-is

    return records


def _merge_applicants(records: list[dict]) -> list[dict]:
    """Merge applicant-section records into current-section records.

    If an applicant unit matches an existing unit, replace the existing
    (typically Vacant) record with the applicant's data.  If no match,
    add the applicant as a new record.  Then strip the internal
    ``_section_type`` tag from all records.
    """
    current = [r for r in records if r.get("_section_type") != "applicants"]
    applicants = [r for r in records if r.get("_section_type") == "applicants"]

    if not applicants:
        for r in current:
            r.pop("_section_type", None)
        return current

    # Build lookup: unit number → index in current list
    unit_index: dict[str, int] = {}
    for i, r in enumerate(current):
        unit_key = str(r.get("unit", "")).strip()
        if unit_key:
            unit_index[unit_key] = i

    for app in applicants:
        app["status"] = "Applicant"
        unit_key = str(app.get("unit", "")).strip()
        if unit_key in unit_index:
            current[unit_index[unit_key]] = app
        else:
            current.append(app)

    for r in current:
        r.pop("_section_type", None)
    return current


def _parse_single_row(ws, col_map: dict,
                      ai_status_map: dict[str, str] | None = None,
                      ai_footer_patterns: list[str] | None = None) -> list[dict]:
    """Extract records from a single-row-per-unit rent roll."""
    cols = col_map["columns"]
    charge_cols = col_map.get("charge_columns") or {}
    data_start = col_map["data_start_row"]
    status_exists = col_map.get("status_column_exists", True)
    section_dividers = col_map.get("section_dividers") or []
    unit_col_idx = _col_index(cols["unit"])

    data_end = _find_data_end(ws, data_start, unit_col_idx, ai_footer_patterns)
    col_map["data_end_row"] = data_end

    section_ranges = _compute_section_ranges(section_dividers, data_start, data_end)

    records = []
    for range_start, range_end, section_type in section_ranges:
        for row_idx in range(range_start, range_end + 1):
            if _is_footer_row(ws, row_idx, ai_footer_patterns):
                break

            unit_val = ws.cell(row=row_idx, column=unit_col_idx).value
            if unit_val is None or str(unit_val).strip() == "":
                continue

            record = {"_source_row": row_idx, "_section_type": section_type}
            for field, col_letter in cols.items():
                if col_letter is None:
                    record[field] = None
                    continue
                col_idx = _col_index(col_letter)
                record[field] = ws.cell(row=row_idx, column=col_idx).value

            # Extract charge columns
            for charge_name, charge_letter in charge_cols.items():
                if charge_letter is None:
                    continue
                col_idx = _col_index(charge_letter)
                val = ws.cell(row=row_idx, column=col_idx).value
                try:
                    record[f"charge_{charge_name}"] = float(val) if val is not None else None
                except (ValueError, TypeError):
                    record[f"charge_{charge_name}"] = None

            # Normalize status
            if status_exists and cols.get("status"):
                record["status"] = normalize_status(record.get("status"), ai_status_map)
            else:
                record["status"] = _infer_status(record, section_type)

            _normalize_record(record)
            records.append(record)

    return _merge_applicants(records)


def _parse_multi_row(ws, col_map: dict,
                     ai_footer_patterns: list[str] | None = None) -> list[dict]:
    """Extract records from a multi-row-per-unit rent roll.

    Each unit spans multiple rows: a header row (with unit info), charge detail
    rows (code + amount), an optional Total row, and a blank separator.
    """
    cols = col_map["columns"]
    data_start = col_map["data_start_row"]
    section_dividers = col_map.get("section_dividers") or []
    multi_cfg = col_map.get("multi_row_config") or {}

    code_col_letter = multi_cfg.get("charge_code_column")
    amount_col_letter = multi_cfg.get("charge_amount_column")
    total_indicator = (multi_cfg.get("total_row_indicator") or "total").lower()
    rent_codes = {c.lower() for c in (multi_cfg.get("rent_charge_codes") or ["rent"])}

    code_col_idx = _col_index(code_col_letter) if code_col_letter else None
    amount_col_idx = _col_index(amount_col_letter) if amount_col_letter else None

    unit_col_idx = _col_index(cols["unit"])

    # Find data_end using footer detection
    data_end = data_start
    last_non_blank = data_start
    blank_streak = 0
    for row_idx in range(data_start, ws.max_row + 1):
        if _is_footer_row(ws, row_idx, ai_footer_patterns):
            break

        has_content = any(
            ws.cell(row=row_idx, column=c).value is not None
            for c in range(1, min(ws.max_column + 1, 15))
        )
        if has_content:
            last_non_blank = row_idx
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak > 10:
                break

    data_end = last_non_blank
    col_map["data_end_row"] = data_end

    section_ranges = _compute_section_ranges(section_dividers, data_start, data_end)

    # Group rows into units, iterating section by section
    records = []
    current_record = None
    charges = {}

    def _finalize_unit(section_type):
        nonlocal current_record, charges
        if current_record is None:
            return
        # Set lease_rent from rent charges
        if "lease_rent" not in current_record or current_record.get("lease_rent") is None:
            rent_total = sum(v for k, v in charges.items() if k in rent_codes and v)
            if rent_total:
                current_record["lease_rent"] = rent_total

        # Set total_billing from Total row or sum of all charges
        if current_record.get("total_billing") is None:
            total = sum(v for v in charges.values() if v)
            if total:
                current_record["total_billing"] = total

        # Add individual charges to record
        for code, amount in charges.items():
            safe_code = code.replace(" ", "_").upper()
            current_record[f"charge_{safe_code}"] = amount

        current_record["_section_type"] = section_type
        current_record["status"] = _infer_status(current_record, section_type)

        _normalize_record(current_record)
        records.append(current_record)
        current_record = None
        charges = {}

    for range_start, range_end, section_type in section_ranges:
        # Finalize any pending unit from a previous section
        _finalize_unit(section_type)

        for row_idx in range(range_start, range_end + 1):
            if _is_footer_row(ws, row_idx, ai_footer_patterns):
                break

            # Check if this is a blank separator row
            has_content = any(
                ws.cell(row=row_idx, column=c).value is not None
                for c in range(1, min(ws.max_column + 1, 15))
            )
            if not has_content:
                _finalize_unit(section_type)
                continue

            unit_val = ws.cell(row=row_idx, column=unit_col_idx).value

            # Check for total row
            if code_col_idx:
                code_val = ws.cell(row=row_idx, column=code_col_idx).value
                if code_val and total_indicator in str(code_val).strip().lower():
                    if amount_col_idx and current_record is not None:
                        total_val = ws.cell(row=row_idx, column=amount_col_idx).value
                        if total_val is not None:
                            try:
                                current_record["total_billing"] = float(total_val)
                            except (ValueError, TypeError):
                                pass
                    continue

            # Is this a new unit row? (unit column has a value)
            if unit_val is not None and str(unit_val).strip() != "":
                _finalize_unit(section_type)

                current_record = {"_source_row": row_idx}
                for field, col_letter in cols.items():
                    if col_letter is None:
                        current_record[field] = None
                        continue
                    col_idx = _col_index(col_letter)
                    current_record[field] = ws.cell(row=row_idx, column=col_idx).value
                charges = {}

                # The first row may also have a charge line
                if code_col_idx and amount_col_idx:
                    code_val = ws.cell(row=row_idx, column=code_col_idx).value
                    amount_val = ws.cell(row=row_idx, column=amount_col_idx).value
                    if code_val and amount_val is not None:
                        code_str = str(code_val).strip().lower()
                        try:
                            charges[code_str] = float(amount_val)
                        except (ValueError, TypeError):
                            pass

            elif current_record is not None:
                # Charge detail row for current unit
                if code_col_idx and amount_col_idx:
                    code_val = ws.cell(row=row_idx, column=code_col_idx).value
                    amount_val = ws.cell(row=row_idx, column=amount_col_idx).value
                    if code_val and amount_val is not None:
                        code_str = str(code_val).strip().lower()
                        if total_indicator not in code_str:
                            try:
                                charges[code_str] = float(amount_val)
                            except (ValueError, TypeError):
                                pass

        # Finalize last unit in this section
        _finalize_unit(section_type)

    return _merge_applicants(records)


def parse_rent_roll(client, file_bytes: bytes, filename: str) -> dict:
    """Parse a rent roll Excel file. Returns dict with keys:
    - 'column_map': the LLM-identified column mapping
    - 'df': pandas DataFrame of extracted unit data
    - 'wb_bytes': the original workbook bytes
    - 'sheet_name': name of the data sheet
    - 'raw_rows': list of dicts with raw row data and row numbers
    - 'as_of_date': datetime or None
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = _pick_data_sheet(wb)
    sheet_name = ws.title

    as_of_date = extract_as_of_date(ws)

    col_map = identify_columns(client, ws)

    # Defensive defaults for new fields
    col_map.setdefault("format", "single_row")
    col_map.setdefault("status_column_exists", True)
    col_map.setdefault("multi_row_config", None)
    col_map.setdefault("section_dividers", [])

    # Extract AI-provided detection hints
    ai_status_map = {k.lower(): v for k, v in col_map.pop("status_mapping", {}).items()}
    ai_footer_patterns = col_map.pop("footer_patterns", [])
    ai_section_keywords = col_map.pop("section_keywords", None)
    # Treat empty dict as None for section keywords
    if not ai_section_keywords:
        ai_section_keywords = None

    # Scan full sheet for section dividers the LLM may have missed
    col_map["section_dividers"] = _scan_section_dividers(
        ws, col_map["data_start_row"], col_map["section_dividers"],
        ai_section_keywords, ai_footer_patterns,
    )

    # Branch on format
    if col_map.get("format") == "multi_row":
        records = _parse_multi_row(ws, col_map, ai_footer_patterns)
    else:
        records = _parse_single_row(ws, col_map, ai_status_map, ai_footer_patterns)

    # Resolve any status values the AI mapping didn't cover
    records = _resolve_unknown_statuses(client, records)

    # Filter out non-unit rows (summary/legend data that slipped through footer detection)
    records = _filter_non_unit_rows(records)

    df = pd.DataFrame(records)

    return {
        "column_map": col_map,
        "df": df,
        "wb_bytes": file_bytes,
        "sheet_name": sheet_name,
        "raw_rows": records,
        "as_of_date": as_of_date,
    }
