"""T12 Output — Multi-sheet Excel with Raw Data, Mapping Detail, Summary, and Checks."""

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.output import (
    FONT_NAME, FONT_SIZE,
    TITLE_FILL, TITLE_FONT, HEADER_FONT, DATA_FONT, BOLD_FONT,
    THIN_SIDE, HAIR_SIDE, NO_SIDE,
    FMT_CURRENCY, FMT_CURRENCY_DEC, FMT_NUMBER,
)
from src.t12_mapper import (
    REVENUE_WATERFALL, OTHER_INCOME,
    CONTROLLABLE_EXPENSES, UNCONTROLLABLE_EXPENSES,
    BELOW_THE_LINE, ALL_CATEGORIES,
)
from src.t12_parser import T12ParseResult

LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CHECK_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CHECK_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _col_letter(idx: int) -> str:
    """1-based index to Excel column letter."""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


# ---------------------------------------------------------------------------
# Sheet 1: Raw Data
# ---------------------------------------------------------------------------

def _write_raw_data_sheet(wb, raw_wb_bytes: bytes, sheet_name: str):
    """Verbatim copy of the uploaded T12."""
    raw_wb = openpyxl.load_workbook(BytesIO(raw_wb_bytes), data_only=True)
    raw_ws = raw_wb[sheet_name]

    ws = wb.create_sheet("Raw Data")
    for row in raw_ws.iter_rows(min_row=1, max_row=raw_ws.max_row, max_col=raw_ws.max_column):
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.number_format:
                new_cell.number_format = cell.number_format

    for merge in raw_ws.merged_cells.ranges:
        ws.merge_cells(str(merge))

    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Mapping Detail
# ---------------------------------------------------------------------------

def _write_mapping_sheet(wb, parse_result: T12ParseResult, mapping: list[dict]):
    """Write the Mapping Detail sheet with formulas referencing Raw Data."""
    ws = wb.create_sheet("Mapping Detail")

    # Column layout: A=Source Row, B=GL Code, C=GL Description, D=Category,
    # E=Section, F-Q=Month 1-12, R=Total, S=Confidence, T=Notes
    headers = [
        "Source Row", "GL Code", "GL Description", "Category", "Section",
    ]
    # Add month headers
    for mh in parse_result.month_headers[:12]:
        headers.append(str(mh))
    # Pad if fewer than 12 months
    while len(headers) < 5 + 12:
        headers.append(f"Month {len(headers) - 4}")
    headers.extend(["Total", "Confidence", "Notes"])

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.alignment = CENTER_ALIGN

    # Build mapping lookup by source_row
    map_lookup = {m["source_row"]: m for m in mapping}

    # Build leaf items lookup
    leaf_lookup = {item.source_row: item for item in parse_result.leaf_items}

    # Month columns in Raw Data
    raw_month_start_letter = _col_letter(parse_result.month_col_start)

    # Write data rows
    for data_row_idx, m in enumerate(mapping, start=2):
        src_row = m["source_row"]
        item = leaf_lookup.get(src_row)
        section_str = " > ".join(item.section_path) if item and item.section_path else ""

        ws.cell(row=data_row_idx, column=1, value=src_row).font = DATA_FONT
        ws.cell(row=data_row_idx, column=2, value=m.get("gl_code", "")).font = DATA_FONT
        ws.cell(row=data_row_idx, column=3, value=m.get("gl_description", "")).font = DATA_FONT
        ws.cell(row=data_row_idx, column=4, value=m["category"]).font = DATA_FONT
        ws.cell(row=data_row_idx, column=5, value=section_str).font = DATA_FONT

        # Month formulas (columns F=6 through Q=17)
        for month_offset in range(12):
            raw_col_idx = parse_result.month_col_start + month_offset
            if raw_col_idx <= parse_result.month_col_end:
                raw_col_letter = _col_letter(raw_col_idx)
                formula = f"='Raw Data'!{raw_col_letter}{src_row}"
            else:
                formula = 0
            cell = ws.cell(row=data_row_idx, column=6 + month_offset, value=formula)
            cell.font = DATA_FONT
            cell.number_format = FMT_CURRENCY

        # Total = SUM of month columns
        total_cell = ws.cell(
            row=data_row_idx, column=18,
            value=f"=SUM(F{data_row_idx}:Q{data_row_idx})",
        )
        total_cell.font = BOLD_FONT
        total_cell.number_format = FMT_CURRENCY

        # Confidence and Notes
        ws.cell(row=data_row_idx, column=19, value=m.get("confidence", "")).font = DATA_FONT
        ws.cell(row=data_row_idx, column=20, value=m.get("notes", "")).font = DATA_FONT

    # Auto-width for key columns
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    for col in range(6, 18):
        ws.column_dimensions[_col_letter(col)].width = 14
    ws.column_dimensions["R"].width = 14
    ws.column_dimensions["S"].width = 12
    ws.column_dimensions["T"].width = 40

    return ws, len(mapping)


# ---------------------------------------------------------------------------
# Sheet 3: T12 Summary
# ---------------------------------------------------------------------------

def _write_summary_sheet(wb, mapping_count: int, unit_count: int | None, total_sf: float | None):
    """Write the T12 Summary sheet with SUMIFS formulas."""
    ws = wb.create_sheet("T12 Summary")

    last_map_row = mapping_count + 1  # 1-based, header is row 1

    # --- Property info row ---
    ws.cell(row=1, column=2, value="T12 Income Statement Summary").font = Font(
        name=FONT_NAME, size=12, bold=True
    )
    ws.cell(row=1, column=5, value="Units:").font = BOLD_FONT
    ws.cell(row=1, column=6, value=unit_count or 0).font = BOLD_FONT
    ws.cell(row=1, column=6).number_format = FMT_NUMBER
    ws.cell(row=1, column=7, value="SF:").font = BOLD_FONT
    ws.cell(row=1, column=8, value=total_sf or 0).font = BOLD_FONT
    ws.cell(row=1, column=8).number_format = FMT_NUMBER

    # --- Headers (row 3) ---
    headers = {2: "Category", 5: "$ Amt.", 6: "$ / Unit", 7: "$ / SF", 8: "Notes"}
    for col, label in headers.items():
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = TITLE_FONT
        cell.fill = TITLE_FILL
        cell.alignment = CENTER_ALIGN

    # Track row positions for formulas
    row_map = {}  # category_name -> excel_row
    current_row = 5  # Start data at row 5

    def _write_category_row(category: str, row: int):
        """Write a SUMIFS row for a category."""
        ws.cell(row=row, column=2, value=category).font = DATA_FONT
        # $ Amt = SUMIFS
        formula = (
            f"=SUMIFS('Mapping Detail'!$R$2:$R${last_map_row},"
            f"'Mapping Detail'!$D$2:$D${last_map_row},\"{category}\")"
        )
        cell_e = ws.cell(row=row, column=5, value=formula)
        cell_e.font = DATA_FONT
        cell_e.number_format = FMT_CURRENCY

        # $/Unit
        cell_f = ws.cell(row=row, column=6, value=f"=IFERROR(E{row}/$F$1,0)")
        cell_f.font = DATA_FONT
        cell_f.number_format = FMT_CURRENCY_DEC

        # $/SF
        cell_g = ws.cell(row=row, column=7, value=f"=IFERROR(E{row}/$H$1,0)")
        cell_g.font = DATA_FONT
        cell_g.number_format = FMT_CURRENCY_DEC

        row_map[category] = row

    def _write_total_row(label: str, formula_str: str, row: int, key: str = None):
        """Write a bold total/calculated row."""
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        cell_e = ws.cell(row=row, column=5, value=formula_str)
        cell_e.font = BOLD_FONT
        cell_e.number_format = FMT_CURRENCY
        cell_e.fill = TOTAL_FILL

        cell_f = ws.cell(row=row, column=6, value=f"=IFERROR(E{row}/$F$1,0)")
        cell_f.font = BOLD_FONT
        cell_f.number_format = FMT_CURRENCY_DEC
        cell_f.fill = TOTAL_FILL

        cell_g = ws.cell(row=row, column=7, value=f"=IFERROR(E{row}/$H$1,0)")
        cell_g.font = BOLD_FONT
        cell_g.number_format = FMT_CURRENCY_DEC
        cell_g.fill = TOTAL_FILL

        if key:
            row_map[key] = row

    def _write_section_header(label: str, row: int):
        """Write a section divider label."""
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=True)
        cell.fill = SECTION_FILL
        for col in range(5, 9):
            ws.cell(row=row, column=col).fill = SECTION_FILL

    # ===== REVENUE WATERFALL =====
    _write_section_header("REVENUE", current_row)
    current_row += 1
    waterfall_start = current_row

    for cat in REVENUE_WATERFALL:
        _write_category_row(cat, current_row)
        current_row += 1

    # Rental Revenue is a sum of the waterfall items above it
    rental_rev_row = row_map.get("Rental Revenue")
    # Overwrite Rental Revenue to be a computed sum (it maps GL items that ARE rental revenue)
    # But also add a "Net" total row
    waterfall_end = current_row - 1
    _write_total_row(
        "Net Rental Revenue",
        f"=SUM(E{waterfall_start}:E{waterfall_end})",
        current_row,
        key="_net_rental_revenue",
    )
    current_row += 2  # blank row

    # ===== OTHER INCOME =====
    _write_section_header("OTHER INCOME", current_row)
    current_row += 1
    other_income_start = current_row

    for cat in OTHER_INCOME:
        _write_category_row(cat, current_row)
        current_row += 1

    other_income_end = current_row - 1
    _write_total_row(
        "Total Other Income",
        f"=SUM(E{other_income_start}:E{other_income_end})",
        current_row,
        key="_total_other_income",
    )
    current_row += 1

    # ===== EGI =====
    nrr_row = row_map["_net_rental_revenue"]
    toi_row = row_map["_total_other_income"]
    _write_total_row(
        "Effective Gross Income",
        f"=E{nrr_row}+E{toi_row}",
        current_row,
        key="_egi",
    )
    current_row += 2

    # ===== CONTROLLABLE EXPENSES =====
    _write_section_header("CONTROLLABLE EXPENSES", current_row)
    current_row += 1
    ctrl_start = current_row

    for cat in CONTROLLABLE_EXPENSES:
        _write_category_row(cat, current_row)
        current_row += 1

    ctrl_end = current_row - 1
    _write_total_row(
        "Total Controllable Expenses",
        f"=SUM(E{ctrl_start}:E{ctrl_end})",
        current_row,
        key="_total_controllable",
    )
    current_row += 2

    # ===== UNCONTROLLABLE EXPENSES =====
    _write_section_header("UNCONTROLLABLE EXPENSES", current_row)
    current_row += 1
    unctrl_start = current_row

    for cat in UNCONTROLLABLE_EXPENSES:
        _write_category_row(cat, current_row)
        current_row += 1

    unctrl_end = current_row - 1
    _write_total_row(
        "Total Uncontrollable Expenses",
        f"=SUM(E{unctrl_start}:E{unctrl_end})",
        current_row,
        key="_total_uncontrollable",
    )
    current_row += 1

    # ===== TOTAL OPEX =====
    tc_row = row_map["_total_controllable"]
    tu_row = row_map["_total_uncontrollable"]
    _write_total_row(
        "Total Operating Expenses",
        f"=E{tc_row}+E{tu_row}",
        current_row,
        key="_total_opex",
    )
    current_row += 1

    # ===== NOI =====
    egi_row = row_map["_egi"]
    opex_row = row_map["_total_opex"]
    _write_total_row(
        "Net Operating Income",
        f"=E{egi_row}+E{opex_row}",
        current_row,
        key="_noi",
    )
    current_row += 2

    # ===== BELOW THE LINE =====
    _write_section_header("BELOW THE LINE", current_row)
    current_row += 1
    btl_start = current_row

    for cat in BELOW_THE_LINE:
        _write_category_row(cat, current_row)
        current_row += 1

    btl_end = current_row - 1
    _write_total_row(
        "Total Below the Line",
        f"=SUM(E{btl_start}:E{btl_end})",
        current_row,
        key="_total_btl",
    )
    current_row += 1

    # ===== CASH FLOW =====
    noi_row = row_map["_noi"]
    btl_row = row_map["_total_btl"]
    _write_total_row(
        "Cash Flow After BTL",
        f"=E{noi_row}+E{btl_row}",
        current_row,
        key="_cash_flow",
    )

    # Column widths
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 30

    return ws, row_map


# ---------------------------------------------------------------------------
# Sheet 4: Checks
# ---------------------------------------------------------------------------

def _write_checks_sheet(
    wb,
    row_map: dict,
    parse_result: T12ParseResult,
):
    """Write reconciliation checks comparing our totals to the T12's own totals."""
    ws = wb.create_sheet("Checks")
    grand_totals = parse_result.grand_totals

    # Header
    check_headers = ["Check", "Our Value", "T12 Source", "Difference", "Status"]
    for col_idx, header in enumerate(check_headers, start=2):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = TITLE_FONT
        cell.fill = TITLE_FILL
        cell.alignment = CENTER_ALIGN

    checks = []

    # EGI check
    egi_row = row_map.get("_egi")
    t12_income = grand_totals.get("total_income")
    if egi_row and t12_income:
        checks.append(("EGI Check", f"='T12 Summary'!E{egi_row}", t12_income))

    # Opex check
    opex_row = row_map.get("_total_opex")
    t12_opex = grand_totals.get("total_opex")
    if opex_row and t12_opex:
        checks.append(("Opex Check", f"='T12 Summary'!E{opex_row}", t12_opex))

    # NOI check
    noi_row = row_map.get("_noi")
    t12_noi = grand_totals.get("noi")
    if noi_row and t12_noi:
        checks.append(("NOI Check", f"='T12 Summary'!E{noi_row}", t12_noi))

    # Net income check (if available)
    cash_flow_row = row_map.get("_cash_flow")
    t12_net = grand_totals.get("net_income")
    if cash_flow_row and t12_net:
        checks.append(("Net Income Check", f"='T12 Summary'!E{cash_flow_row}", t12_net))

    for i, (label, our_formula, t12_item) in enumerate(checks, start=2):
        ws.cell(row=i, column=2, value=label).font = BOLD_FONT

        # Our value (formula referencing Summary)
        cell_c = ws.cell(row=i, column=3, value=our_formula)
        cell_c.font = DATA_FONT
        cell_c.number_format = FMT_CURRENCY

        # T12 source (formula referencing Raw Data)
        raw_total_col_letter = _col_letter(parse_result.total_col)
        t12_formula = f"='Raw Data'!{raw_total_col_letter}{t12_item.source_row}"
        cell_d = ws.cell(row=i, column=4, value=t12_formula)
        cell_d.font = DATA_FONT
        cell_d.number_format = FMT_CURRENCY

        # Difference
        cell_e = ws.cell(row=i, column=5, value=f"=C{i}-D{i}")
        cell_e.font = DATA_FONT
        cell_e.number_format = FMT_CURRENCY

        # Status
        cell_f = ws.cell(
            row=i, column=6,
            value=f'=IF(ABS(E{i})<1,"PASS","FAIL")',
        )
        cell_f.font = BOLD_FONT

    # Conditional formatting (manual since openpyxl CF is complex)
    # Just set a note
    if checks:
        ws.cell(row=len(checks) + 3, column=2, value="Note: PASS means difference < $1").font = Font(
            name=FONT_NAME, size=FONT_SIZE, italic=True, color="666666"
        )

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12

    return ws


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_t12_output(
    parse_result: T12ParseResult,
    mapping: list[dict],
    raw_wb_bytes: bytes,
    unit_count: int | None = None,
    total_sf: float | None = None,
) -> bytes:
    """Generate the T12 output workbook and return as bytes."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet 1: Raw Data
    _write_raw_data_sheet(wb, raw_wb_bytes, parse_result.sheet_name)

    # Sheet 2: Mapping Detail
    _, mapping_count = _write_mapping_sheet(wb, parse_result, mapping)

    # Sheet 3: T12 Summary
    _, row_map = _write_summary_sheet(wb, mapping_count, unit_count, total_sf)

    # Sheet 4: Checks
    _write_checks_sheet(wb, row_map, parse_result)

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
